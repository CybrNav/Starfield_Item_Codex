##!/usr/bin/env python3
# -*- coding: utf-8 -*-
# StarfieldItemCodex.py - Starfield item database tool
# (formerly AstralCodex.py, renamed bc it outgrew just astral ui)
# needs: customtkinter, openpyxl
# drop next to FormID_List.xlsx and run it
# Current version: 1.10

from __future__ import annotations

import os
import re
import sys
import threading
import tkinter as tk
from collections import namedtuple
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Optional, Tuple

import customtkinter as ctk
import openpyxl

# --- appearance (must happen before any CTk widget) ---
# clam theme looks least garbage in dark mode
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

## constants ##
APP_TITLE    = "Starfield Item Codex"
APP_W, APP_H = 1660, 920
MIN_W, MIN_H = 1430, 805
XLSX_NAME    = "FormID_List.xlsx"
REQ_HEADERS  = {"Category", "SubCategory", "Name", "FormID", "EditorID", "Source"}
PAD          = 10

# treeview dark palette
TV_BG     = "#2b2b2b"
TV_FG     = "#DCE4EE"
TV_ALT    = "#313131"
TV_SEL    = "#3B8ED0"
TV_HDR_BG = "#1e1e1e"
TV_HDR_FG = "#DCE4EE"

# ESMs whose FormID prefix may vary with load order (not hardcoded)
VARIABLE_ESMS = frozenset({
    "SFBGS00D.esm", "SFBGS050.esm",
    "sfta01.esm", "sfta02.esm", "sfta03.esm",
    "sfta04.esm", "sfta05.esm", "sfta06.esm",
})

# column definitions for each tab's treeview
# first column is a tiny checkbox indicator (added to right panel or not)
LOOKUP_COLS   = ("added", "category", "subcategory", "name", "form_id", "editor_id", "source")
LOOKUP_HEADS  = ("", "Category", "SubCategory", "Name", "FormID", "EditorID", "Source")
LOOKUP_WIDTHS = (30, 110, 130, 220, 110, 185, 110)

# subcategory builder right panel
BLDR_COLS   = ("name", "form_id")
BLDR_HEADS  = ("Name", "FormID")
BLDR_WIDTHS = (270, 130)

# batch creator right panel
BATCH_COLS   = ("name", "form_id", "qty")
BATCH_HEADS  = ("Name", "FormID", "Qty")
BATCH_WIDTHS = (240, 130, 60)

Item = namedtuple("Item", ["category", "subcategory", "name", "form_id", "editor_id", "source"])

# i18n
_current_lang: str = "English"

# display names for the language dropdown: "english name - native name"
# english stays as just "english" since it's the same in both
LANG_DISPLAY_NAMES: Dict[str, str] = {
    "English":            "English",
    "German":             "German - Deutsch",
    "Spanish":            "Spanish - Espanol",
    "French":             "French - Francais",
    "Italian":            "Italian - Italiano",
    "Japanese":           "Japanese - 日本語",
    "Polish":             "Polish - Polski",
    "Portuguese-BR":      "Portuguese-BR - Portugues",
    "Chinese-Simplified": "Chinese-Simplified - 简体中文",
}
# reverse lookup: display name -> sheet name
LANG_FROM_DISPLAY: Dict[str, str] = {v: k for k, v in LANG_DISPLAY_NAMES.items()}

# tab display names per language (sheet name -> dict of tab_id -> label)
# yeah... i know
TAB_IDS = ("tab_lookup", "tab_batch", "tab_builder")
TAB_NAMES_EN = {
    "tab_lookup":  "🔍 Search",
    "tab_batch":   "🎮 Batch File Creator",
    "tab_builder": "🏗 AstralUI Subcategory Builder",
}
TAB_NAMES = {
    "English":            {
        "tab_lookup":  "🔍 Search",
        "tab_batch":   "🎮 Batch File Creator",
        "tab_builder": "🏗 AstralUI Subcategory Builder",
    },
    "German":             {
        "tab_lookup":  "🔍 Suche",
        "tab_batch":   "🎮 Batch-Datei Ersteller",
        "tab_builder": "🏗 AstralUI Unterkategorie",
    },
    "Spanish":            {
        "tab_lookup":  "🔍 Buscar",
        "tab_batch":   "🎮 Creador de Batch",
        "tab_builder": "🏗 AstralUI Subcategorias",
    },
    "French":             {
        "tab_lookup":  "🔍 Recherche",
        "tab_batch":   "🎮 Createur de Batch",
        "tab_builder": "🏗 AstralUI Sous-categories",
    },
    "Italian":            {
        "tab_lookup":  "🔍 Ricerca",
        "tab_batch":   "🎮 Creatore Batch",
        "tab_builder": "🏗 AstralUI Sottocategorie",
    },
    "Japanese":           {
        "tab_lookup":  "🔍 検索",
        "tab_batch":   "🎮 バッチ作成",
        "tab_builder": "🏗 AstralUI カテゴリ",
    },
    "Polish":             {
        "tab_lookup":  "🔍 Szukaj",
        "tab_batch":   "🎮 Kreator Batch",
        "tab_builder": "🏗 AstralUI Kategorie",
    },
    "Portuguese-BR":      {
        "tab_lookup":  "🔍 Buscar",
        "tab_batch":   "🎮 Criador de Batch",
        "tab_builder": "🏗 AstralUI Subcategorias",
    },
    "Chinese-Simplified": {
        "tab_lookup":  "🔍 搜索",
        "tab_batch":   "🎮 批处理创建",
        "tab_builder": "🏗 AstralUI 分类",
    },
}

# fmt: off
_EN = {
    # description banner
    "desc_lookup":    "Search and browse the complete Starfield item database",
    "desc_builder":   "Create custom subcategories for your AstralUI INI",
    "desc_batch":     "Generate console batch files to spawn items in-game",
    # labels
    "language":       "Language:",
    "search":         "Search:",
    "category":       "Category:",
    "subcategory":    "SubCategory:",
    "all_cat":        "All Categories",
    "all_subcat":     "All SubCategories",
    "subcat_name":    "Subcategory Name:",
    "your_subcat":    "Your Subcategory",
    "ini_preview":    "INI Preview:",
    "batch_items":    "Batch Items",
    "set_all_qty":    "Set All Qty:",
    "set_sel_qty":    "Set Selected Qty:",
    "entire_cat":     "Entire Category:",
    "entire_subcat":  "Entire SubCategory:",
    # column headers
    "col_category":   "Category",
    "col_subcategory":"SubCategory",
    "col_name":       "Name",
    "col_formid":     "FormID",
    "col_editorid":   "EditorID",
    "col_source":     "Source",
    "col_qty":        "Qty",
    # buttons
    "add_selected":   "Add Selected",
    "add_all_vis":    "Add All Visible",
    "add_category":   "Add Category",
    "add_subcat":     "Add SubCategory",
    "remove_sel":     "Remove Selected",
    "clear_all":      "Clear All",
    "copy_clipboard": "Copy to Clipboard",
    "save_batch":     "Save Batch File",
    "copy_commands":  "Copy Commands",
    "apply":          "Apply",
    # placeholders
    "search_ph":      "Name, FormID, EditorID...",
    "name_ph":        "Enter a name... (spaces to underscores, no = ; [ ])",
    # context menu
    "ctx_formid":     "Copy FormID",
    "ctx_console":    "Copy Console Command (player.additem)",
    "ctx_row":        "Copy Row",
    # status / counts
    "items_shown":    "{n} items shown",
    "one_item_shown": "1 item shown",
    "no_match":       "No items match.",
    "n_items":        "{n} items",
    "one_item":       "1 item",
    "copied_fid":     "Copied FormID: {fid}",
    "copied_cmd":     "Copied: {cmd}",
    "copied_row":     "Copied row to clipboard",
    "added_to_sub":   "Added {n} item(s) to subcategory",
    "added_to_batch": "Added {n} item(s) to batch",
    "no_sel_search":  "No items selected in the search results",
    "no_vis":         "No visible items to add",
    "sel_cat_first":  "Select a category first",
    "sel_cat_sub":    "Select a category and subcategory first",
    "removed_sub":    "Removed {n} item(s) from subcategory",
    "removed_batch":  "Removed {n} item(s) from batch",
    "cleared_sub":    "Subcategory list cleared",
    "cleared_batch":  "Batch list cleared",
    "ini_copied":     "INI line copied to clipboard",
    "batch_saved":    "Batch file saved: {f}",
    "cmds_copied":    "Copied {n} console command(s) to clipboard",
    "qty_set":        "Set all quantities to {n}",
    "qty_set_sel":    "Set {c} selected to qty {n}",
    "no_sel_for_qty": "No items selected in the batch list",
    "add_to_enable":  "Add items to enable output",
    "n_ready":        "{n} item(s) ready",
    "name_autofixed": "Subcategory name was auto-fixed (spaces to underscores, stripped invalid chars)",
    "no_sel_sub":     "No items selected in your subcategory list",
    "no_sel_batch":   "No items selected in the batch list",
    "removed_item":   "Removed item from subcategory",
    # dialogs
    "dlg_clear_all":  "Clear All",
    "dlg_clear_sub":  "Remove all {n} items from your subcategory list?",
    "dlg_clear_batch":"Remove all {n} items from the batch list?",
    "dlg_unsaved":    "Unsaved Work",
    "dlg_switch_tab": "Your Subcategory Builder list has items.\n\nSwitch tabs anyway? (Your list will remain - it is NOT cleared.)",
    "dlg_exit":       "The following lists have items:\n\n{detail}\n\nExit anyway? (All unsaved work will be lost.)",
    "dlg_sub_list":   "Subcategory Builder list",
    "dlg_batch_list": "Batch Creator list",
    "dlg_not_found":  "Workbook Not Found",
    "dlg_not_found_msg": "'{name}' was not found next to this application.\n\nPlease locate the file manually.",
    "dlg_no_file":    "No File Selected",
    "dlg_no_file_msg":"No workbook was selected.\n\nStarfield Item Codex cannot run without the FormID database.\nPlace the .xlsx next to the application and restart.",
    "dlg_load_err":   "Load Error",
    "dlg_save_err":   "Save Error",
    "dlg_write_fail": "Could not write file:\n{e}",
    "clipboard_err":  "Clipboard error - could not copy",
    "invalid_qty":    "Invalid quantity - enter a whole number >= 1",
    # loading screen
    "loading_title":  "Starfield Item Codex",
    "loading_msg":    "Loading Starfield item database...",
    # batch file name
    "batch_name":     "Batch File Name:",
    "batch_name_ph":  "Enter a filename for your batch...",
    # subcategory copy button
    "copy_subcat":    "Copy Subcategory To Clipboard",
    # cross-tab copy
    "copy_from_batch":"Copy from Batch",
    "copy_from_bldr": "Copy from Builder",
    "copied_from_batch": "Copied {n} item(s) from Batch list",
    "copied_from_bldr":  "Copied {n} item(s) from Builder list",
    "nothing_to_copy":"Nothing to copy - other list is empty",
    # preview placeholder
    "preview_ph":     "(enter a name and add items to see the INI line)",
    # paste existing subcategory
    "paste_existing":  "Paste Existing Subcategory",
    "paste_dlg_title": "Paste Existing Subcategory",
    "paste_dlg_instr": "Paste one or more INI subcategory lines below:\ne.g. My_Weapons=0x002C5884,0x002995A3,...",
    "paste_dlg_add":   "Add",
    "paste_dlg_cancel": "Cancel",
    "paste_imported":  "Imported {n} item(s) from pasted subcategory ({u} unknown)",
    "paste_empty":     "Nothing to import - paste a subcategory line first",
    "paste_no_fids":   "No valid FormIDs found in pasted text",
    # hardcoded filter
    "hardcoded_only": "Only Show Hardcoded FormIDs",
    # loaded status
    "loaded_status":  "Loaded {total} items across {langs} language(s)",
    "load_failed":    "Load failed - {e}",
    "starting":       "Starting up...",
}

TRANSLATIONS: Dict[str, Dict[str, str]] = {
    "English": _EN,
    "German": {
        "desc_lookup":    "Durchsuche die komplette Starfield-Gegenstands-Datenbank",
        "desc_builder":   "Erstelle eigene Unterkategorien fur deine AstralUI-INI",
        "desc_batch":     "Erstelle Konsolen-Batch-Dateien zum Spawnen von Gegenstanden",
        "language":       "Sprache:",
        "search":         "Suche:",
        "category":       "Kategorie:",
        "subcategory":    "Unterkategorie:",
        "all_cat":        "Alle Kategorien",
        "all_subcat":     "Alle Unterkategorien",
        "subcat_name":    "Name der Unterkategorie:",
        "your_subcat":    "Deine Unterkategorie",
        "ini_preview":    "INI-Vorschau:",
        "batch_items":    "Batch-Gegenstande",
        "set_all_qty":    "Alle Mengen setzen:",
        "set_sel_qty":    "Auswahl Menge:",
        "entire_cat":     "Ganze Kategorie:",
        "entire_subcat":  "Ganze Unterkategorie:",
        "col_category":   "Kategorie",
        "col_subcategory":"Unterkategorie",
        "col_name":       "Name",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Quelle",
        "col_qty":        "Anz.",
        "add_selected":   "Auswahl hinzufugen",
        "add_all_vis":    "Alle sichtbaren hinzufugen",
        "add_category":   "Kategorie hinzufugen",
        "add_subcat":     "Unterkategorie hinzufugen",
        "remove_sel":     "Auswahl entfernen",
        "clear_all":      "Alle leeren",
        "copy_clipboard": "In Zwischenablage kopieren",
        "save_batch":     "Batch-Datei speichern",
        "copy_commands":  "Befehle kopieren",
        "apply":          "Anwenden",
        "search_ph":      "Name, FormID, EditorID...",
        "name_ph":        "Name eingeben... (Leerzeichen werden zu Unterstrichen)",
        "ctx_formid":     "FormID kopieren",
        "ctx_console":    "Konsolenbefehl kopieren (player.additem)",
        "ctx_row":        "Zeile kopieren",
        "items_shown":    "{n} Gegenstande angezeigt",
        "one_item_shown": "1 Gegenstand angezeigt",
        "no_match":       "Keine Treffer.",
        "n_items":        "{n} Gegenstande",
        "one_item":       "1 Gegenstand",
        "copied_fid":     "FormID kopiert: {fid}",
        "copied_cmd":     "Kopiert: {cmd}",
        "copied_row":     "Zeile in Zwischenablage kopiert",
        "added_to_sub":   "{n} Gegenstand/Gegenstande zur Unterkategorie hinzugefugt",
        "added_to_batch": "{n} Gegenstand/Gegenstande zum Batch hinzugefugt",
        "no_sel_search":  "Keine Gegenstande in den Suchergebnissen ausgewahlt",
        "no_vis":         "Keine sichtbaren Gegenstande zum Hinzufugen",
        "sel_cat_first":  "Wahle zuerst eine Kategorie",
        "sel_cat_sub":    "Wahle zuerst eine Kategorie und Unterkategorie",
        "removed_sub":    "{n} Gegenstand/Gegenstande aus Unterkategorie entfernt",
        "removed_batch":  "{n} Gegenstand/Gegenstande aus Batch entfernt",
        "cleared_sub":    "Unterkategorie-Liste geleert",
        "cleared_batch":  "Batch-Liste geleert",
        "ini_copied":     "INI-Zeile in Zwischenablage kopiert",
        "batch_saved":    "Batch-Datei gespeichert: {f}",
        "cmds_copied":    "{n} Konsolenbefehl(e) in Zwischenablage kopiert",
        "qty_set":        "Alle Mengen auf {n} gesetzt",
        "qty_set_sel":    "{c} ausgewahlte auf Menge {n} gesetzt",
        "no_sel_for_qty": "Keine Gegenstande in der Batch-Liste ausgewahlt",
        "add_to_enable":  "Gegenstande hinzufugen um Ausgabe zu aktivieren",
        "n_ready":        "{n} Gegenstand/Gegenstande bereit",
        "name_autofixed": "Name wurde automatisch korrigiert (Leerzeichen zu Unterstrichen)",
        "no_sel_sub":     "Keine Gegenstande in deiner Unterkategorie-Liste ausgewahlt",
        "no_sel_batch":   "Keine Gegenstande in der Batch-Liste ausgewahlt",
        "removed_item":   "Gegenstand aus Unterkategorie entfernt",
        "dlg_clear_all":  "Alle leeren",
        "dlg_clear_sub":  "Alle {n} Gegenstande aus deiner Unterkategorie-Liste entfernen?",
        "dlg_clear_batch":"Alle {n} Gegenstande aus der Batch-Liste entfernen?",
        "dlg_unsaved":    "Ungespeicherte Arbeit",
        "dlg_switch_tab": "Deine Unterkategorie-Liste enthalt Gegenstande.\n\nTrotzdem wechseln? (Die Liste bleibt erhalten.)",
        "dlg_exit":       "Folgende Listen enthalten Gegenstande:\n\n{detail}\n\nTrotzdem beenden? (Ungespeicherte Arbeit geht verloren.)",
        "dlg_sub_list":   "Unterkategorie-Liste",
        "dlg_batch_list": "Batch-Liste",
        "dlg_not_found":  "Arbeitsmappe nicht gefunden",
        "dlg_not_found_msg": "'{name}' wurde nicht neben dieser Anwendung gefunden.\n\nBitte manuell suchen.",
        "dlg_no_file":    "Keine Datei ausgewahlt",
        "dlg_no_file_msg":"Keine Arbeitsmappe ausgewahlt.\n\nStarfield Item Codex benotigt die FormID-Datenbank.\nLege die .xlsx neben die Anwendung und starte neu.",
        "dlg_load_err":   "Ladefehler",
        "dlg_save_err":   "Speicherfehler",
        "dlg_write_fail": "Datei konnte nicht geschrieben werden:\n{e}",
        "clipboard_err":  "Zwischenablage-Fehler",
        "invalid_qty":    "Ungultige Menge - ganze Zahl >= 1 eingeben",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Lade Starfield-Gegenstands-Datenbank...",
        "preview_ph":     "(Name eingeben und Gegenstande hinzufugen)",
        "paste_existing":  "Bestehende Unterkategorie einfugen",
        "paste_dlg_title": "Bestehende Unterkategorie einfugen",
        "paste_dlg_instr": "Fuge eine oder mehrere INI-Unterkategorie-Zeilen ein:\nz.B. Meine_Waffen=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Hinzufugen",
        "paste_dlg_cancel": "Abbrechen",
        "paste_imported":  "{n} Gegenstand/Gegenstande aus eingefugter Unterkategorie importiert ({u} unbekannt)",
        "paste_empty":     "Nichts zu importieren - fuge zuerst eine Unterkategorie-Zeile ein",
        "paste_no_fids":   "Keine gultigen FormIDs im eingefugten Text gefunden",
        "hardcoded_only": "Nur feste FormIDs",
        "loaded_status":  "{total} Gegenstande in {langs} Sprache(n) geladen",
        "load_failed":    "Laden fehlgeschlagen - {e}",
        "starting":       "Starte...",
    },
    "Spanish": {
        "desc_lookup":    "Busca y explora la base de datos completa de objetos de Starfield",
        "desc_builder":   "Crea subcategorias personalizadas para tu INI de AstralUI",
        "desc_batch":     "Genera archivos batch de consola para spawnear objetos en el juego",
        "language":       "Idioma:",
        "search":         "Buscar:",
        "category":       "Categoria:",
        "subcategory":    "Subcategoria:",
        "all_cat":        "Todas las Categorias",
        "all_subcat":     "Todas las Subcategorias",
        "subcat_name":    "Nombre de Subcategoria:",
        "your_subcat":    "Tu Subcategoria",
        "ini_preview":    "Vista previa INI:",
        "batch_items":    "Objetos del Batch",
        "set_all_qty":    "Cantidad para todos:",
        "set_sel_qty":    "Cantidad seleccion:",
        "entire_cat":     "Categoria completa:",
        "entire_subcat":  "Subcategoria completa:",
        "col_category":   "Categoria",
        "col_subcategory":"Subcategoria",
        "col_name":       "Nombre",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Origen",
        "col_qty":        "Cant.",
        "add_selected":   "Agregar seleccion",
        "add_all_vis":    "Agregar todos visibles",
        "add_category":   "Agregar categoria",
        "add_subcat":     "Agregar subcategoria",
        "remove_sel":     "Quitar seleccion",
        "clear_all":      "Vaciar todo",
        "copy_clipboard": "Copiar al portapapeles",
        "save_batch":     "Guardar archivo batch",
        "copy_commands":  "Copiar comandos",
        "apply":          "Aplicar",
        "search_ph":      "Nombre, FormID, EditorID...",
        "name_ph":        "Ingresa un nombre... (espacios a guiones bajos)",
        "ctx_formid":     "Copiar FormID",
        "ctx_console":    "Copiar comando de consola (player.additem)",
        "ctx_row":        "Copiar fila",
        "items_shown":    "{n} objetos mostrados",
        "one_item_shown": "1 objeto mostrado",
        "no_match":       "Sin resultados.",
        "n_items":        "{n} objetos",
        "one_item":       "1 objeto",
        "copied_fid":     "FormID copiado: {fid}",
        "copied_cmd":     "Copiado: {cmd}",
        "copied_row":     "Fila copiada al portapapeles",
        "added_to_sub":   "{n} objeto(s) agregado(s) a la subcategoria",
        "added_to_batch": "{n} objeto(s) agregado(s) al batch",
        "no_sel_search":  "No hay objetos seleccionados en los resultados",
        "no_vis":         "No hay objetos visibles para agregar",
        "sel_cat_first":  "Selecciona una categoria primero",
        "sel_cat_sub":    "Selecciona una categoria y subcategoria primero",
        "removed_sub":    "{n} objeto(s) eliminado(s) de la subcategoria",
        "removed_batch":  "{n} objeto(s) eliminado(s) del batch",
        "cleared_sub":    "Lista de subcategoria vaciada",
        "cleared_batch":  "Lista de batch vaciada",
        "ini_copied":     "Linea INI copiada al portapapeles",
        "batch_saved":    "Archivo batch guardado: {f}",
        "cmds_copied":    "{n} comando(s) de consola copiado(s)",
        "qty_set":        "Todas las cantidades establecidas a {n}",
        "qty_set_sel":    "{c} seleccionados a cantidad {n}",
        "no_sel_for_qty": "No hay objetos seleccionados en la lista",
        "add_to_enable":  "Agrega objetos para habilitar la salida",
        "n_ready":        "{n} objeto(s) listo(s)",
        "name_autofixed": "Nombre corregido automaticamente (espacios a guiones bajos)",
        "no_sel_sub":     "No hay objetos seleccionados en tu lista de subcategoria",
        "no_sel_batch":   "No hay objetos seleccionados en la lista de batch",
        "removed_item":   "Objeto eliminado de la subcategoria",
        "dlg_clear_all":  "Vaciar todo",
        "dlg_clear_sub":  "Eliminar los {n} objetos de tu lista de subcategoria?",
        "dlg_clear_batch":"Eliminar los {n} objetos de la lista de batch?",
        "dlg_unsaved":    "Trabajo sin guardar",
        "dlg_switch_tab": "Tu lista de subcategoria tiene objetos.\n\nCambiar de pestana? (La lista se mantiene.)",
        "dlg_exit":       "Las siguientes listas tienen objetos:\n\n{detail}\n\nSalir de todas formas? (El trabajo no guardado se perdera.)",
        "dlg_sub_list":   "Lista de subcategoria",
        "dlg_batch_list": "Lista de batch",
        "dlg_not_found":  "Libro de trabajo no encontrado",
        "dlg_not_found_msg": "'{name}' no se encontro junto a esta aplicacion.\n\nPor favor, buscalo manualmente.",
        "dlg_no_file":    "Ningun archivo seleccionado",
        "dlg_no_file_msg":"No se selecciono ningun libro de trabajo.\n\nStarfield Item Codex necesita la base de datos de FormID.\nColoca el .xlsx junto a la aplicacion y reinicia.",
        "dlg_load_err":   "Error de carga",
        "dlg_save_err":   "Error al guardar",
        "dlg_write_fail": "No se pudo escribir el archivo:\n{e}",
        "clipboard_err":  "Error del portapapeles",
        "invalid_qty":    "Cantidad invalida - ingresa un numero entero >= 1",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Cargando base de datos de objetos de Starfield...",
        "preview_ph":     "(ingresa un nombre y agrega objetos para ver la linea INI)",
        "paste_existing":  "Pegar subcategoria existente",
        "paste_dlg_title": "Pegar subcategoria existente",
        "paste_dlg_instr": "Pega una o mas lineas de subcategoria INI abajo:\nej. Mis_Armas=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Agregar",
        "paste_dlg_cancel": "Cancelar",
        "paste_imported":  "{n} objeto(s) importado(s) de subcategoria pegada ({u} desconocido(s))",
        "paste_empty":     "Nada que importar - pega primero una linea de subcategoria",
        "paste_no_fids":   "No se encontraron FormIDs validos en el texto pegado",
        "hardcoded_only": "Solo FormIDs fijos",
        "loaded_status":  "{total} objetos cargados en {langs} idioma(s)",
        "load_failed":    "Error de carga - {e}",
        "starting":       "Iniciando...",
    },
    "French": {
        "desc_lookup":    "Recherchez et parcourez la base de donnees complete des objets Starfield",
        "desc_builder":   "Creez des sous-categories personnalisees pour votre INI AstralUI",
        "desc_batch":     "Generez des fichiers batch console pour faire apparaitre des objets en jeu",
        "language":       "Langue :",
        "search":         "Recherche :",
        "category":       "Categorie :",
        "subcategory":    "Sous-categorie :",
        "all_cat":        "Toutes les categories",
        "all_subcat":     "Toutes les sous-categories",
        "subcat_name":    "Nom de sous-categorie :",
        "your_subcat":    "Votre sous-categorie",
        "ini_preview":    "Apercu INI :",
        "batch_items":    "Objets du batch",
        "set_all_qty":    "Quantite pour tous :",
        "set_sel_qty":    "Quantite selection :",
        "entire_cat":     "Categorie entiere :",
        "entire_subcat":  "Sous-categorie entiere :",
        "col_category":   "Categorie",
        "col_subcategory":"Sous-categorie",
        "col_name":       "Nom",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Source",
        "col_qty":        "Qte",
        "add_selected":   "Ajouter la selection",
        "add_all_vis":    "Ajouter tous les visibles",
        "add_category":   "Ajouter la categorie",
        "add_subcat":     "Ajouter la sous-categorie",
        "remove_sel":     "Retirer la selection",
        "clear_all":      "Tout effacer",
        "copy_clipboard": "Copier dans le presse-papiers",
        "save_batch":     "Enregistrer le fichier batch",
        "copy_commands":  "Copier les commandes",
        "apply":          "Appliquer",
        "search_ph":      "Nom, FormID, EditorID...",
        "name_ph":        "Entrez un nom... (espaces en tirets bas)",
        "ctx_formid":     "Copier le FormID",
        "ctx_console":    "Copier la commande console (player.additem)",
        "ctx_row":        "Copier la ligne",
        "items_shown":    "{n} objets affiches",
        "one_item_shown": "1 objet affiche",
        "no_match":       "Aucun resultat.",
        "n_items":        "{n} objets",
        "one_item":       "1 objet",
        "copied_fid":     "FormID copie : {fid}",
        "copied_cmd":     "Copie : {cmd}",
        "copied_row":     "Ligne copiee dans le presse-papiers",
        "added_to_sub":   "{n} objet(s) ajoute(s) a la sous-categorie",
        "added_to_batch": "{n} objet(s) ajoute(s) au batch",
        "no_sel_search":  "Aucun objet selectionne dans les resultats",
        "no_vis":         "Aucun objet visible a ajouter",
        "sel_cat_first":  "Selectionnez d'abord une categorie",
        "sel_cat_sub":    "Selectionnez une categorie et une sous-categorie",
        "removed_sub":    "{n} objet(s) retire(s) de la sous-categorie",
        "removed_batch":  "{n} objet(s) retire(s) du batch",
        "cleared_sub":    "Liste de sous-categorie videe",
        "cleared_batch":  "Liste de batch videe",
        "ini_copied":     "Ligne INI copiee dans le presse-papiers",
        "batch_saved":    "Fichier batch enregistre : {f}",
        "cmds_copied":    "{n} commande(s) console copiee(s)",
        "qty_set":        "Toutes les quantites definies a {n}",
        "qty_set_sel":    "{c} selectionnes a quantite {n}",
        "no_sel_for_qty": "Aucun objet selectionne dans la liste",
        "add_to_enable":  "Ajoutez des objets pour activer la sortie",
        "n_ready":        "{n} objet(s) pret(s)",
        "name_autofixed": "Nom corrige automatiquement (espaces en tirets bas)",
        "no_sel_sub":     "Aucun objet selectionne dans votre liste",
        "no_sel_batch":   "Aucun objet selectionne dans la liste batch",
        "removed_item":   "Objet retire de la sous-categorie",
        "dlg_clear_all":  "Tout effacer",
        "dlg_clear_sub":  "Retirer les {n} objets de votre liste de sous-categorie ?",
        "dlg_clear_batch":"Retirer les {n} objets de la liste batch ?",
        "dlg_unsaved":    "Travail non enregistre",
        "dlg_switch_tab": "Votre liste de sous-categorie contient des objets.\n\nChanger d'onglet ? (La liste est conservee.)",
        "dlg_exit":       "Les listes suivantes contiennent des objets :\n\n{detail}\n\nQuitter quand meme ? (Le travail non enregistre sera perdu.)",
        "dlg_sub_list":   "Liste de sous-categorie",
        "dlg_batch_list": "Liste batch",
        "dlg_not_found":  "Classeur introuvable",
        "dlg_not_found_msg": "'{name}' n'a pas ete trouve a cote de cette application.\n\nVeuillez le localiser manuellement.",
        "dlg_no_file":    "Aucun fichier selectionne",
        "dlg_no_file_msg":"Aucun classeur selectionne.\n\nStarfield Item Codex a besoin de la base de donnees FormID.\nPlacez le .xlsx a cote de l'application et redemarrez.",
        "dlg_load_err":   "Erreur de chargement",
        "dlg_save_err":   "Erreur d'enregistrement",
        "dlg_write_fail": "Impossible d'ecrire le fichier :\n{e}",
        "clipboard_err":  "Erreur du presse-papiers",
        "invalid_qty":    "Quantite invalide - entrez un nombre entier >= 1",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Chargement de la base de donnees Starfield...",
        "preview_ph":     "(entrez un nom et ajoutez des objets pour voir la ligne INI)",
        "paste_existing":  "Coller une sous-categorie existante",
        "paste_dlg_title": "Coller une sous-categorie existante",
        "paste_dlg_instr": "Collez une ou plusieurs lignes de sous-categorie INI ci-dessous :\nex. Mes_Armes=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Ajouter",
        "paste_dlg_cancel": "Annuler",
        "paste_imported":  "{n} objet(s) importe(s) depuis la sous-categorie collee ({u} inconnu(s))",
        "paste_empty":     "Rien a importer - collez d'abord une ligne de sous-categorie",
        "paste_no_fids":   "Aucun FormID valide trouve dans le texte colle",
        "hardcoded_only": "FormIDs fixes uniquement",
        "loaded_status":  "{total} objets charges dans {langs} langue(s)",
        "load_failed":    "Echec du chargement - {e}",
        "starting":       "Demarrage...",
    },
    "Italian": {
        "desc_lookup":    "Cerca e sfoglia il database completo degli oggetti di Starfield",
        "desc_builder":   "Crea sottocategorie personalizzate per il tuo INI AstralUI",
        "desc_batch":     "Genera file batch console per spawnare oggetti nel gioco",
        "language":       "Lingua:",
        "search":         "Cerca:",
        "category":       "Categoria:",
        "subcategory":    "Sottocategoria:",
        "all_cat":        "Tutte le Categorie",
        "all_subcat":     "Tutte le Sottocategorie",
        "subcat_name":    "Nome Sottocategoria:",
        "your_subcat":    "La tua Sottocategoria",
        "ini_preview":    "Anteprima INI:",
        "batch_items":    "Oggetti Batch",
        "set_all_qty":    "Imposta tutte le quantita:",
        "set_sel_qty":    "Quantita selezione:",
        "entire_cat":     "Categoria intera:",
        "entire_subcat":  "Sottocategoria intera:",
        "col_category":   "Categoria",
        "col_subcategory":"Sottocategoria",
        "col_name":       "Nome",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Fonte",
        "col_qty":        "Qta",
        "add_selected":   "Aggiungi selezionati",
        "add_all_vis":    "Aggiungi tutti visibili",
        "add_category":   "Aggiungi categoria",
        "add_subcat":     "Aggiungi sottocategoria",
        "remove_sel":     "Rimuovi selezionati",
        "clear_all":      "Svuota tutto",
        "copy_clipboard": "Copia negli appunti",
        "save_batch":     "Salva file batch",
        "copy_commands":  "Copia comandi",
        "apply":          "Applica",
        "search_ph":      "Nome, FormID, EditorID...",
        "name_ph":        "Inserisci un nome... (spazi in trattini bassi)",
        "ctx_formid":     "Copia FormID",
        "ctx_console":    "Copia comando console (player.additem)",
        "ctx_row":        "Copia riga",
        "items_shown":    "{n} oggetti mostrati",
        "one_item_shown": "1 oggetto mostrato",
        "no_match":       "Nessun risultato.",
        "n_items":        "{n} oggetti",
        "one_item":       "1 oggetto",
        "copied_fid":     "FormID copiato: {fid}",
        "copied_cmd":     "Copiato: {cmd}",
        "copied_row":     "Riga copiata negli appunti",
        "added_to_sub":   "{n} oggetto/i aggiunto/i alla sottocategoria",
        "added_to_batch": "{n} oggetto/i aggiunto/i al batch",
        "no_sel_search":  "Nessun oggetto selezionato nei risultati",
        "no_vis":         "Nessun oggetto visibile da aggiungere",
        "sel_cat_first":  "Seleziona prima una categoria",
        "sel_cat_sub":    "Seleziona prima una categoria e sottocategoria",
        "removed_sub":    "{n} oggetto/i rimosso/i dalla sottocategoria",
        "removed_batch":  "{n} oggetto/i rimosso/i dal batch",
        "cleared_sub":    "Lista sottocategoria svuotata",
        "cleared_batch":  "Lista batch svuotata",
        "ini_copied":     "Riga INI copiata negli appunti",
        "batch_saved":    "File batch salvato: {f}",
        "cmds_copied":    "{n} comando/i console copiato/i",
        "qty_set":        "Tutte le quantita impostate a {n}",
        "qty_set_sel":    "{c} selezionati a quantita {n}",
        "no_sel_for_qty": "Nessun oggetto selezionato nella lista",
        "add_to_enable":  "Aggiungi oggetti per abilitare l'output",
        "n_ready":        "{n} oggetto/i pronto/i",
        "name_autofixed": "Nome corretto automaticamente (spazi in trattini bassi)",
        "no_sel_sub":     "Nessun oggetto selezionato nella tua lista",
        "no_sel_batch":   "Nessun oggetto selezionato nella lista batch",
        "removed_item":   "Oggetto rimosso dalla sottocategoria",
        "dlg_clear_all":  "Svuota tutto",
        "dlg_clear_sub":  "Rimuovere tutti i {n} oggetti dalla lista sottocategoria?",
        "dlg_clear_batch":"Rimuovere tutti i {n} oggetti dalla lista batch?",
        "dlg_unsaved":    "Lavoro non salvato",
        "dlg_switch_tab": "La tua lista sottocategoria contiene oggetti.\n\nCambiare scheda? (La lista rimane.)",
        "dlg_exit":       "Le seguenti liste contengono oggetti:\n\n{detail}\n\nUscire comunque? (Il lavoro non salvato sara perso.)",
        "dlg_sub_list":   "Lista sottocategoria",
        "dlg_batch_list": "Lista batch",
        "dlg_not_found":  "Cartella di lavoro non trovata",
        "dlg_not_found_msg": "'{name}' non trovato accanto a questa applicazione.\n\nCercalo manualmente.",
        "dlg_no_file":    "Nessun file selezionato",
        "dlg_no_file_msg":"Nessuna cartella di lavoro selezionata.\n\nStarfield Item Codex necessita del database FormID.\nPosiziona il .xlsx accanto all'applicazione e riavvia.",
        "dlg_load_err":   "Errore di caricamento",
        "dlg_save_err":   "Errore di salvataggio",
        "dlg_write_fail": "Impossibile scrivere il file:\n{e}",
        "clipboard_err":  "Errore degli appunti",
        "invalid_qty":    "Quantita non valida - inserisci un numero intero >= 1",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Caricamento database oggetti Starfield...",
        "preview_ph":     "(inserisci un nome e aggiungi oggetti per vedere la riga INI)",
        "paste_existing":  "Incolla sottocategoria esistente",
        "paste_dlg_title": "Incolla sottocategoria esistente",
        "paste_dlg_instr": "Incolla una o piu righe di sottocategoria INI qui sotto:\nes. Le_Mie_Armi=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Aggiungi",
        "paste_dlg_cancel": "Annulla",
        "paste_imported":  "{n} oggetto/i importato/i dalla sottocategoria incollata ({u} sconosciuto/i)",
        "paste_empty":     "Niente da importare - incolla prima una riga di sottocategoria",
        "paste_no_fids":   "Nessun FormID valido trovato nel testo incollato",
        "hardcoded_only": "Solo FormID fissi",
        "loaded_status":  "{total} oggetti caricati in {langs} lingua/e",
        "load_failed":    "Caricamento fallito - {e}",
        "starting":       "Avvio...",
    },
    "Japanese": {
        "desc_lookup":    "Starfieldの全アイテムデータベースを検索・閲覧",
        "desc_builder":   "AstralUI INI用のカスタムサブカテゴリを作成",
        "desc_batch":     "ゲーム内でアイテムをスポーンするコンソールバッチファイルを生成",
        "language":       "言語:",
        "search":         "検索:",
        "category":       "カテゴリ:",
        "subcategory":    "サブカテゴリ:",
        "all_cat":        "全てのカテゴリ",
        "all_subcat":     "全てのサブカテゴリ",
        "subcat_name":    "サブカテゴリ名:",
        "your_subcat":    "あなたのサブカテゴリ",
        "ini_preview":    "INIプレビュー:",
        "batch_items":    "バッチアイテム",
        "set_all_qty":    "全数量を設定:",
        "set_sel_qty":    "選択数量:",
        "entire_cat":     "カテゴリ全体:",
        "entire_subcat":  "サブカテゴリ全体:",
        "col_category":   "カテゴリ",
        "col_subcategory":"サブカテゴリ",
        "col_name":       "名前",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "ソース",
        "col_qty":        "数量",
        "add_selected":   "選択を追加",
        "add_all_vis":    "表示中を全て追加",
        "add_category":   "カテゴリを追加",
        "add_subcat":     "サブカテゴリを追加",
        "remove_sel":     "選択を削除",
        "clear_all":      "全てクリア",
        "copy_clipboard": "クリップボードにコピー",
        "save_batch":     "バッチファイルを保存",
        "copy_commands":  "コマンドをコピー",
        "apply":          "適用",
        "search_ph":      "名前、FormID、EditorID...",
        "name_ph":        "名前を入力... (スペースはアンダースコアに変換)",
        "ctx_formid":     "FormIDをコピー",
        "ctx_console":    "コンソールコマンドをコピー (player.additem)",
        "ctx_row":        "行をコピー",
        "items_shown":    "{n}件表示中",
        "one_item_shown": "1件表示中",
        "no_match":       "該当なし",
        "n_items":        "{n}件",
        "one_item":       "1件",
        "copied_fid":     "FormIDをコピーしました: {fid}",
        "copied_cmd":     "コピーしました: {cmd}",
        "copied_row":     "行をクリップボードにコピーしました",
        "added_to_sub":   "{n}件をサブカテゴリに追加しました",
        "added_to_batch": "{n}件をバッチに追加しました",
        "no_sel_search":  "検索結果で選択されていません",
        "no_vis":         "追加できるアイテムがありません",
        "sel_cat_first":  "先にカテゴリを選択してください",
        "sel_cat_sub":    "先にカテゴリとサブカテゴリを選択してください",
        "removed_sub":    "{n}件をサブカテゴリから削除しました",
        "removed_batch":  "{n}件をバッチから削除しました",
        "cleared_sub":    "サブカテゴリリストをクリアしました",
        "cleared_batch":  "バッチリストをクリアしました",
        "ini_copied":     "INI行をクリップボードにコピーしました",
        "batch_saved":    "バッチファイルを保存しました: {f}",
        "cmds_copied":    "{n}件のコンソールコマンドをコピーしました",
        "qty_set":        "全数量を{n}に設定しました",
        "qty_set_sel":    "{c}個を数量{n}に設定",
        "no_sel_for_qty": "バッチリストでアイテムが選択されていません",
        "add_to_enable":  "アイテムを追加して出力を有効にしてください",
        "n_ready":        "{n}件準備完了",
        "name_autofixed": "名前を自動修正しました (スペースをアンダースコアに変換)",
        "no_sel_sub":     "サブカテゴリリストで選択されていません",
        "no_sel_batch":   "バッチリストで選択されていません",
        "removed_item":   "アイテムをサブカテゴリから削除しました",
        "dlg_clear_all":  "全てクリア",
        "dlg_clear_sub":  "サブカテゴリリストの{n}件を全て削除しますか?",
        "dlg_clear_batch":"{n}件のバッチアイテムを全て削除しますか?",
        "dlg_unsaved":    "未保存の作業",
        "dlg_switch_tab": "サブカテゴリリストにアイテムがあります。\n\nタブを切り替えますか? (リストは保持されます)",
        "dlg_exit":       "以下のリストにアイテムがあります:\n\n{detail}\n\n終了しますか? (未保存の作業は失われます)",
        "dlg_sub_list":   "サブカテゴリリスト",
        "dlg_batch_list": "バッチリスト",
        "dlg_not_found":  "ワークブックが見つかりません",
        "dlg_not_found_msg": "'{name}'がこのアプリケーションの隣に見つかりませんでした。\n\n手動で探してください。",
        "dlg_no_file":    "ファイルが選択されていません",
        "dlg_no_file_msg":"ワークブックが選択されていません。\n\nStarfield Item CodexにはFormIDデータベースが必要です。\n.xlsxをアプリケーションの隣に置いて再起動してください。",
        "dlg_load_err":   "読み込みエラー",
        "dlg_save_err":   "保存エラー",
        "dlg_write_fail": "ファイルを書き込めませんでした:\n{e}",
        "clipboard_err":  "クリップボードエラー",
        "invalid_qty":    "無効な数量 - 1以上の整数を入力してください",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Starfieldアイテムデータベースを読み込み中...",
        "preview_ph":     "(名前を入力してアイテムを追加するとINI行が表示されます)",
        "paste_existing":  "既存サブカテゴリを貼り付け",
        "paste_dlg_title": "既存サブカテゴリを貼り付け",
        "paste_dlg_instr": "INIサブカテゴリ行を以下に貼り付けてください:\n例: My_Weapons=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "追加",
        "paste_dlg_cancel": "キャンセル",
        "paste_imported":  "貼り付けたサブカテゴリから{n}件をインポートしました ({u}件不明)",
        "paste_empty":     "インポートするものがありません - 先にサブカテゴリ行を貼り付けてください",
        "paste_no_fids":   "貼り付けたテキストに有効なFormIDが見つかりません",
        "hardcoded_only": "固定FormIDのみ",
        "loaded_status":  "{langs}言語で{total}件のアイテムを読み込みました",
        "load_failed":    "読み込み失敗 - {e}",
        "starting":       "起動中...",
    },
    "Polish": {
        "desc_lookup":    "Przeszukuj i przegladaj kompletna baze przedmiotow Starfield",
        "desc_builder":   "Tworzenie wlasnych podkategorii dla INI AstralUI",
        "desc_batch":     "Generowanie plikow batch konsoli do spawnowania przedmiotow w grze",
        "language":       "Jezyk:",
        "search":         "Szukaj:",
        "category":       "Kategoria:",
        "subcategory":    "Podkategoria:",
        "all_cat":        "Wszystkie Kategorie",
        "all_subcat":     "Wszystkie Podkategorie",
        "subcat_name":    "Nazwa Podkategorii:",
        "your_subcat":    "Twoja Podkategoria",
        "ini_preview":    "Podglad INI:",
        "batch_items":    "Przedmioty Batch",
        "set_all_qty":    "Ustaw wszystkie ilosci:",
        "set_sel_qty":    "Ustaw ilosc wybranych:",
        "entire_cat":     "Cala Kategoria:",
        "entire_subcat":  "Cala Podkategoria:",
        "col_category":   "Kategoria",
        "col_subcategory":"Podkategoria",
        "col_name":       "Nazwa",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Zrodlo",
        "col_qty":        "Ilosc",
        "add_selected":   "Dodaj zaznaczone",
        "add_all_vis":    "Dodaj wszystkie widoczne",
        "add_category":   "Dodaj kategorie",
        "add_subcat":     "Dodaj podkategorie",
        "remove_sel":     "Usun zaznaczone",
        "clear_all":      "Wyczysc wszystko",
        "copy_clipboard": "Kopiuj do schowka",
        "save_batch":     "Zapisz plik batch",
        "copy_commands":  "Kopiuj komendy",
        "apply":          "Zastosuj",
        "search_ph":      "Nazwa, FormID, EditorID...",
        "name_ph":        "Wpisz nazwe... (spacje zamienione na podkreslenia)",
        "ctx_formid":     "Kopiuj FormID",
        "ctx_console":    "Kopiuj komende konsoli (player.additem)",
        "ctx_row":        "Kopiuj wiersz",
        "items_shown":    "{n} przedmiotow wyswietlonych",
        "one_item_shown": "1 przedmiot wyswietlony",
        "no_match":       "Brak wynikow.",
        "n_items":        "{n} przedmiotow",
        "one_item":       "1 przedmiot",
        "copied_fid":     "Skopiowano FormID: {fid}",
        "copied_cmd":     "Skopiowano: {cmd}",
        "copied_row":     "Wiersz skopiowany do schowka",
        "added_to_sub":   "Dodano {n} przedmiot(ow) do podkategorii",
        "added_to_batch": "Dodano {n} przedmiot(ow) do batcha",
        "no_sel_search":  "Nie zaznaczono przedmiotow w wynikach wyszukiwania",
        "no_vis":         "Brak widocznych przedmiotow do dodania",
        "sel_cat_first":  "Najpierw wybierz kategorie",
        "sel_cat_sub":    "Najpierw wybierz kategorie i podkategorie",
        "removed_sub":    "Usunieto {n} przedmiot(ow) z podkategorii",
        "removed_batch":  "Usunieto {n} przedmiot(ow) z batcha",
        "cleared_sub":    "Lista podkategorii wyczyszczona",
        "cleared_batch":  "Lista batcha wyczyszczona",
        "ini_copied":     "Linia INI skopiowana do schowka",
        "batch_saved":    "Plik batch zapisany: {f}",
        "cmds_copied":    "Skopiowano {n} komend(e) konsoli do schowka",
        "qty_set":        "Wszystkie ilosci ustawione na {n}",
        "qty_set_sel":    "{c} wybranych ustawionych na ilosc {n}",
        "no_sel_for_qty": "Nie wybrano przedmiotow na liscie batch",
        "add_to_enable":  "Dodaj przedmioty aby aktywowac wyjscie",
        "n_ready":        "{n} przedmiot(ow) gotowych",
        "name_autofixed": "Nazwa automatycznie poprawiona (spacje na podkreslenia)",
        "no_sel_sub":     "Nie zaznaczono przedmiotow na liscie podkategorii",
        "no_sel_batch":   "Nie zaznaczono przedmiotow na liscie batcha",
        "removed_item":   "Usunieto przedmiot z podkategorii",
        "dlg_clear_all":  "Wyczysc wszystko",
        "dlg_clear_sub":  "Usunac wszystkie {n} przedmiotow z listy podkategorii?",
        "dlg_clear_batch":"Usunac wszystkie {n} przedmiotow z listy batcha?",
        "dlg_unsaved":    "Niezapisana praca",
        "dlg_switch_tab": "Lista podkategorii zawiera przedmioty.\n\nZmienic zakladke? (Lista zostanie zachowana.)",
        "dlg_exit":       "Nastepujace listy zawieraja przedmioty:\n\n{detail}\n\nWyjsc mimo to? (Niezapisana praca zostanie utracona.)",
        "dlg_sub_list":   "Lista podkategorii",
        "dlg_batch_list": "Lista batcha",
        "dlg_not_found":  "Nie znaleziono skoroszytu",
        "dlg_not_found_msg": "'{name}' nie znaleziono obok tej aplikacji.\n\nZlokalizuj plik recznie.",
        "dlg_no_file":    "Nie wybrano pliku",
        "dlg_no_file_msg":"Nie wybrano skoroszytu.\n\nStarfield Item Codex wymaga bazy danych FormID.\nUmiesc .xlsx obok aplikacji i uruchom ponownie.",
        "dlg_load_err":   "Blad ladowania",
        "dlg_save_err":   "Blad zapisu",
        "dlg_write_fail": "Nie mozna zapisac pliku:\n{e}",
        "clipboard_err":  "Blad schowka",
        "invalid_qty":    "Nieprawidlowa ilosc - wpisz liczbe calkowita >= 1",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Ladowanie bazy danych przedmiotow Starfield...",
        "preview_ph":     "(wpisz nazwe i dodaj przedmioty aby zobaczyc linie INI)",
        "paste_existing":  "Wklej istniejaca podkategorie",
        "paste_dlg_title": "Wklej istniejaca podkategorie",
        "paste_dlg_instr": "Wklej jedna lub wiecej linii podkategorii INI ponizej:\nnp. Moje_Bronie=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Dodaj",
        "paste_dlg_cancel": "Anuluj",
        "paste_imported":  "Zaimportowano {n} przedmiot(ow) z wklejonej podkategorii ({u} nieznanych)",
        "paste_empty":     "Nic do importu - najpierw wklej linie podkategorii",
        "paste_no_fids":   "Nie znaleziono prawidlowych FormID w wklejonym tekscie",
        "hardcoded_only": "Tylko stale FormID",
        "loaded_status":  "Zaladowano {total} przedmiotow w {langs} jezyku/ach",
        "load_failed":    "Ladowanie nieudane - {e}",
        "starting":       "Uruchamianie...",
    },
    "Portuguese-BR": {
        "desc_lookup":    "Pesquise e navegue pelo banco de dados completo de itens do Starfield",
        "desc_builder":   "Crie subcategorias personalizadas para o seu INI do AstralUI",
        "desc_batch":     "Gere arquivos batch de console para spawnar itens no jogo",
        "language":       "Idioma:",
        "search":         "Buscar:",
        "category":       "Categoria:",
        "subcategory":    "Subcategoria:",
        "all_cat":        "Todas as Categorias",
        "all_subcat":     "Todas as Subcategorias",
        "subcat_name":    "Nome da Subcategoria:",
        "your_subcat":    "Sua Subcategoria",
        "ini_preview":    "Previa INI:",
        "batch_items":    "Itens do Batch",
        "set_all_qty":    "Definir todas as quantidades:",
        "set_sel_qty":    "Qtd. selecionados:",
        "entire_cat":     "Categoria inteira:",
        "entire_subcat":  "Subcategoria inteira:",
        "col_category":   "Categoria",
        "col_subcategory":"Subcategoria",
        "col_name":       "Nome",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "Fonte",
        "col_qty":        "Qtd",
        "add_selected":   "Adicionar selecionados",
        "add_all_vis":    "Adicionar todos visiveis",
        "add_category":   "Adicionar categoria",
        "add_subcat":     "Adicionar subcategoria",
        "remove_sel":     "Remover selecionados",
        "clear_all":      "Limpar tudo",
        "copy_clipboard": "Copiar para area de transferencia",
        "save_batch":     "Salvar arquivo batch",
        "copy_commands":  "Copiar comandos",
        "apply":          "Aplicar",
        "search_ph":      "Nome, FormID, EditorID...",
        "name_ph":        "Digite um nome... (espacos viram underscores)",
        "ctx_formid":     "Copiar FormID",
        "ctx_console":    "Copiar comando de console (player.additem)",
        "ctx_row":        "Copiar linha",
        "items_shown":    "{n} itens exibidos",
        "one_item_shown": "1 item exibido",
        "no_match":       "Sem resultados.",
        "n_items":        "{n} itens",
        "one_item":       "1 item",
        "copied_fid":     "FormID copiado: {fid}",
        "copied_cmd":     "Copiado: {cmd}",
        "copied_row":     "Linha copiada para area de transferencia",
        "added_to_sub":   "{n} item(ns) adicionado(s) a subcategoria",
        "added_to_batch": "{n} item(ns) adicionado(s) ao batch",
        "no_sel_search":  "Nenhum item selecionado nos resultados",
        "no_vis":         "Nenhum item visivel para adicionar",
        "sel_cat_first":  "Selecione uma categoria primeiro",
        "sel_cat_sub":    "Selecione uma categoria e subcategoria primeiro",
        "removed_sub":    "{n} item(ns) removido(s) da subcategoria",
        "removed_batch":  "{n} item(ns) removido(s) do batch",
        "cleared_sub":    "Lista de subcategoria limpa",
        "cleared_batch":  "Lista de batch limpa",
        "ini_copied":     "Linha INI copiada para area de transferencia",
        "batch_saved":    "Arquivo batch salvo: {f}",
        "cmds_copied":    "{n} comando(s) de console copiado(s)",
        "qty_set":        "Todas as quantidades definidas para {n}",
        "qty_set_sel":    "{c} selecionados para quantidade {n}",
        "no_sel_for_qty": "Nenhum item selecionado na lista batch",
        "add_to_enable":  "Adicione itens para habilitar a saida",
        "n_ready":        "{n} item(ns) pronto(s)",
        "name_autofixed": "Nome corrigido automaticamente (espacos para underscores)",
        "no_sel_sub":     "Nenhum item selecionado na sua lista de subcategoria",
        "no_sel_batch":   "Nenhum item selecionado na lista de batch",
        "removed_item":   "Item removido da subcategoria",
        "dlg_clear_all":  "Limpar tudo",
        "dlg_clear_sub":  "Remover todos os {n} itens da sua lista de subcategoria?",
        "dlg_clear_batch":"Remover todos os {n} itens da lista de batch?",
        "dlg_unsaved":    "Trabalho nao salvo",
        "dlg_switch_tab": "Sua lista de subcategoria tem itens.\n\nTrocar de aba? (A lista sera mantida.)",
        "dlg_exit":       "As seguintes listas tem itens:\n\n{detail}\n\nSair mesmo assim? (O trabalho nao salvo sera perdido.)",
        "dlg_sub_list":   "Lista de subcategoria",
        "dlg_batch_list": "Lista de batch",
        "dlg_not_found":  "Pasta de trabalho nao encontrada",
        "dlg_not_found_msg": "'{name}' nao foi encontrado ao lado desta aplicacao.\n\nPor favor, localize manualmente.",
        "dlg_no_file":    "Nenhum arquivo selecionado",
        "dlg_no_file_msg":"Nenhuma pasta de trabalho selecionada.\n\nStarfield Item Codex precisa do banco de dados FormID.\nColoque o .xlsx ao lado da aplicacao e reinicie.",
        "dlg_load_err":   "Erro de carregamento",
        "dlg_save_err":   "Erro ao salvar",
        "dlg_write_fail": "Nao foi possivel gravar o arquivo:\n{e}",
        "clipboard_err":  "Erro da area de transferencia",
        "invalid_qty":    "Quantidade invalida - insira um numero inteiro >= 1",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "Carregando banco de dados de itens do Starfield...",
        "preview_ph":     "(digite um nome e adicione itens para ver a linha INI)",
        "paste_existing":  "Colar subcategoria existente",
        "paste_dlg_title": "Colar subcategoria existente",
        "paste_dlg_instr": "Cole uma ou mais linhas de subcategoria INI abaixo:\nex. Minhas_Armas=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "Adicionar",
        "paste_dlg_cancel": "Cancelar",
        "paste_imported":  "{n} item(ns) importado(s) da subcategoria colada ({u} desconhecido(s))",
        "paste_empty":     "Nada para importar - cole primeiro uma linha de subcategoria",
        "paste_no_fids":   "Nenhum FormID valido encontrado no texto colado",
        "hardcoded_only": "Apenas FormIDs fixos",
        "loaded_status":  "{total} itens carregados em {langs} idioma(s)",
        "load_failed":    "Falha no carregamento - {e}",
        "starting":       "Iniciando...",
    },
    "Chinese-Simplified": {
        "desc_lookup":    "搜索和浏览完整的星空物品数据库",
        "desc_builder":   "为AstralUI INI创建自定义子分类",
        "desc_batch":     "生成控制台批处理文件以在游戏中生成物品",
        "language":       "语言:",
        "search":         "搜索:",
        "category":       "分类:",
        "subcategory":    "子分类:",
        "all_cat":        "所有分类",
        "all_subcat":     "所有子分类",
        "subcat_name":    "子分类名称:",
        "your_subcat":    "你的子分类",
        "ini_preview":    "INI预览:",
        "batch_items":    "批处理物品",
        "set_all_qty":    "设置所有数量:",
        "set_sel_qty":    "设置选中数量:",
        "entire_cat":     "整个分类:",
        "entire_subcat":  "整个子分类:",
        "col_category":   "分类",
        "col_subcategory":"子分类",
        "col_name":       "名称",
        "col_formid":     "FormID",
        "col_editorid":   "EditorID",
        "col_source":     "来源",
        "col_qty":        "数量",
        "add_selected":   "添加选中项",
        "add_all_vis":    "添加所有可见项",
        "add_category":   "添加分类",
        "add_subcat":     "添加子分类",
        "remove_sel":     "移除选中项",
        "clear_all":      "全部清除",
        "copy_clipboard": "复制到剪贴板",
        "save_batch":     "保存批处理文件",
        "copy_commands":  "复制命令",
        "apply":          "应用",
        "search_ph":      "名称、FormID、EditorID...",
        "name_ph":        "输入名称... (空格转为下划线)",
        "ctx_formid":     "复制FormID",
        "ctx_console":    "复制控制台命令 (player.additem)",
        "ctx_row":        "复制行",
        "items_shown":    "显示{n}个物品",
        "one_item_shown": "显示1个物品",
        "no_match":       "无匹配结果",
        "n_items":        "{n}个物品",
        "one_item":       "1个物品",
        "copied_fid":     "已复制FormID: {fid}",
        "copied_cmd":     "已复制: {cmd}",
        "copied_row":     "已复制行到剪贴板",
        "added_to_sub":   "已添加{n}个物品到子分类",
        "added_to_batch": "已添加{n}个物品到批处理",
        "no_sel_search":  "搜索结果中未选择任何物品",
        "no_vis":         "没有可添加的可见物品",
        "sel_cat_first":  "请先选择一个分类",
        "sel_cat_sub":    "请先选择分类和子分类",
        "removed_sub":    "已从子分类中移除{n}个物品",
        "removed_batch":  "已从批处理中移除{n}个物品",
        "cleared_sub":    "子分类列表已清除",
        "cleared_batch":  "批处理列表已清除",
        "ini_copied":     "INI行已复制到剪贴板",
        "batch_saved":    "批处理文件已保存: {f}",
        "cmds_copied":    "已复制{n}条控制台命令到剪贴板",
        "qty_set":        "所有数量已设置为{n}",
        "qty_set_sel":    "已将{c}个选中项设置为数量{n}",
        "no_sel_for_qty": "未在批处理列表中选择物品",
        "add_to_enable":  "添加物品以启用输出",
        "n_ready":        "{n}个物品就绪",
        "name_autofixed": "名称已自动修正 (空格转为下划线)",
        "no_sel_sub":     "子分类列表中未选择任何物品",
        "no_sel_batch":   "批处理列表中未选择任何物品",
        "removed_item":   "已从子分类中移除物品",
        "dlg_clear_all":  "全部清除",
        "dlg_clear_sub":  "从子分类列表中移除全部{n}个物品?",
        "dlg_clear_batch":"从批处理列表中移除全部{n}个物品?",
        "dlg_unsaved":    "未保存的工作",
        "dlg_switch_tab": "你的子分类列表中有物品。\n\n切换标签页? (列表将保留)",
        "dlg_exit":       "以下列表中有物品:\n\n{detail}\n\n确定退出? (未保存的工作将丢失)",
        "dlg_sub_list":   "子分类列表",
        "dlg_batch_list": "批处理列表",
        "dlg_not_found":  "未找到工作簿",
        "dlg_not_found_msg": "在此应用程序旁未找到'{name}'。\n\n请手动定位文件。",
        "dlg_no_file":    "未选择文件",
        "dlg_no_file_msg":"未选择工作簿。\n\nStarfield Item Codex需要FormID数据库。\n请将.xlsx放在应用程序旁并重新启动。",
        "dlg_load_err":   "加载错误",
        "dlg_save_err":   "保存错误",
        "dlg_write_fail": "无法写入文件:\n{e}",
        "clipboard_err":  "剪贴板错误",
        "invalid_qty":    "无效数量 - 请输入>=1的整数",
        "loading_title":  "Starfield Item Codex",
        "loading_msg":    "正在加载星空物品数据库...",
        "preview_ph":     "(输入名称并添加物品以查看INI行)",
        "paste_existing":  "粘贴现有子分类",
        "paste_dlg_title": "粘贴现有子分类",
        "paste_dlg_instr": "在下方粘贴一行或多行INI子分类行:\n例如: My_Weapons=0x002C5884,0x002995A3,...",
        "paste_dlg_add":   "添加",
        "paste_dlg_cancel": "取消",
        "paste_imported":  "从粘贴的子分类导入了{n}个物品 ({u}个未知)",
        "paste_empty":     "没有可导入的内容 - 请先粘贴子分类行",
        "paste_no_fids":   "在粘贴的文本中未找到有效的FormID",
        "hardcoded_only": "仅显示固定FormID",
        "loaded_status":  "已加载{langs}种语言的{total}个物品",
        "load_failed":    "加载失败 - {e}",
        "starting":       "启动中...",
    },
}
# fmt: on

# fill missing languages with english fallback
for _lang_name in ("German", "Spanish", "French", "Italian", "Japanese",
                   "Polish", "Portuguese-BR", "Chinese-Simplified"):
    _tbl = TRANSLATIONS.setdefault(_lang_name, {})
    for _k, _v in _EN.items():
        _tbl.setdefault(_k, _v)


def t(key: str, lang: str | None = None, **kwargs: Any) -> str:
    """Translate key to current language. kwargs for {placeholder} substitution."""
    global _current_lang
    if lang is None:
        lang = _current_lang
    tbl = TRANSLATIONS.get(lang, _EN)
    s = tbl.get(key, _EN.get(key, key))
    if kwargs:
        try:
            s = s.format(**kwargs)
        except (KeyError, IndexError):
            pass
    return s



def setup_treeview_style() -> None:
    s = ttk.Style()
    try:
        s.theme_use("clam")
    except tk.TclError:
        pass

    s.configure("Dark.Treeview",
                background=TV_BG, foreground=TV_FG, fieldbackground=TV_BG,
                borderwidth=0, rowheight=24)
    s.configure("Dark.Treeview.Heading",
                background=TV_HDR_BG, foreground=TV_HDR_FG,
                relief="flat", borderwidth=1)
    s.map("Dark.Treeview",
          background=[("selected", TV_SEL)],
          foreground=[("selected", "#ffffff")])
    s.map("Dark.Treeview.Heading",
          background=[("active", "#252525")])

    for orient in ("Vertical", "Horizontal"):
        s.configure(f"Dark.{orient}.TScrollbar",
                    background="#3d3d3d", troughcolor="#222222",
                    arrowcolor="#888888", borderwidth=0)


def sanitize_ini_key(raw: str) -> Tuple[str, bool]:
    """Spaces to underscores, strip = ; [ ] chars. Returns (sanitized, changed)."""
    s = raw.replace(" ", "_")
    s = re.sub(r"[=;\[\]\r\n]", "", s)
    return s, s != raw


def clipboard_set(widget: tk.Misc, text: str) -> bool:
    try:
        widget.clipboard_clear()
        widget.clipboard_append(text)
        widget.update()
        return True
    except tk.TclError:
        return False


def find_xlsx() -> Optional[str]:
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(base, XLSX_NAME)
    return candidate if os.path.isfile(candidate) else None


def _get_cell(row_tuple: tuple, col_idx: Dict[str, int], key: str) -> str:
    i = col_idx.get(key, -1)
    if i < 0 or i >= len(row_tuple):
        return ""
    v = row_tuple[i]
    return str(v).strip() if v is not None else ""


def load_xlsx(path: str) -> Tuple[Dict[str, List[Item]], List[str], Optional[str]]:
    """Load all sheets that have the required columns. Returns (data, sheet_names, error)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        return {}, [], f"File not found:\n{path}"
    except Exception as exc:
        return {}, [], f"Cannot open workbook:\n{exc}"

    data: Dict[str, List[Item]] = {}
    names: List[str] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        try:
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            continue
        if not rows:
            continue

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        if not REQ_HEADERS.issubset(set(header)):
            continue  # skip sheets that don't have our columns

        col_idx: Dict[str, int] = {h: i for i, h in enumerate(header)}
        items: List[Item] = []

        for row in rows[1:]:
            if not any(row):
                continue
            fid = _get_cell(row, col_idx, "FormID")
            if not fid:
                continue
            items.append(Item(
                category   =_get_cell(row, col_idx, "Category"),
                subcategory=_get_cell(row, col_idx, "SubCategory"),
                name       =_get_cell(row, col_idx, "Name"),
                form_id    =fid,
                editor_id  =_get_cell(row, col_idx, "EditorID"),
                source     =_get_cell(row, col_idx, "Source"),
            ))

        if items:
            data[sname] = items
            names.append(sname)

    wb.close()

    # FIXME: this breaks if the xlsx has merged cells
    if not data:
        return {}, [], (
            "No valid worksheets found in the workbook.\n"
            f"Each sheet must contain: {', '.join(sorted(REQ_HEADERS))}"
        )
    return data, names, None


# --- DarkTreeview ---

class DarkTreeview(tk.Frame):
    """Dark-themed ttk.Treeview with scrollbars, sortable columns, and row stripes."""

    def __init__(self,
                 parent: tk.Widget,
                 columns: Tuple[str, ...],
                 headings: Tuple[str, ...],
                 widths: Tuple[int, ...],
                 selectmode: str = "browse",
                 **_kw: Any) -> None:
        super().__init__(parent, bg=TV_BG, bd=0, highlightthickness=0)

        self.cols = columns
        self._sort_col: Optional[str] = None
        self._sort_rev: bool = False

        self.tree = ttk.Treeview(
            self, columns=columns, show="headings",
            style="Dark.Treeview", selectmode=selectmode
        )
        vsb = ttk.Scrollbar(self, orient="vertical",   command=self.tree.yview,
                             style="Dark.Vertical.TScrollbar")
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview,
                             style="Dark.Horizontal.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        for col, head, w in zip(columns, headings, widths):
            self.tree.heading(col, text=head,
                              command=lambda c=col: self._sort_by(c))
            # checkbox column stays tiny and non-resizable
            if col == "added":
                self.tree.column(col, width=w, minwidth=w, stretch=False, anchor="center")
            else:
                self.tree.column(col, width=w, minwidth=50)

        self.tree.tag_configure("even", background=TV_BG)
        self.tree.tag_configure("odd",  background=TV_ALT)
        self.tree.tag_configure("added_even", background="#1e3a1e")
        self.tree.tag_configure("added_odd",  background="#243d24")

    # sorting

    def _sort_by(self, col: str) -> None:
        # don't sort the checkbox column
        if col == "added":
            return
        if self._sort_col == col:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col
            self._sort_rev = False

        kids = list(self.tree.get_children(""))
        kids.sort(key=lambda k: str(self.tree.set(k, col)).lower(),
                  reverse=self._sort_rev)
        for i, k in enumerate(kids):
            self.tree.move(k, "", i)

        arrow = " ▼" if self._sort_rev else " ▲"
        for c in self.cols:
            txt = self.tree.heading(c)["text"].rstrip(" ▲▼")
            self.tree.heading(c, text=txt + (arrow if c == col else ""))
        self._restripe()

    def _restripe(self) -> None:
        for i, k in enumerate(self.tree.get_children("")):
            self.tree.item(k, tags=("odd" if i % 2 else "even",))


    def clear(self) -> None:
        self.tree.delete(*self.tree.get_children(""))

    def populate(self, rows: List[Tuple], added_fids: set | None = None,
                  fid_col_idx: int = -1) -> None:
        self.clear()
        for i, row in enumerate(rows):
            is_added = (added_fids is not None and fid_col_idx >= 0
                        and str(row[fid_col_idx]) in added_fids)
            if is_added:
                tag = "added_odd" if i % 2 else "added_even"
            else:
                tag = "odd" if i % 2 else "even"
            self.tree.insert("", "end", values=row, tags=(tag,))

    def append_row(self, row: Tuple) -> str:
        i = self.count()
        return self.tree.insert("", "end", values=row,
                                tags=("odd" if i % 2 else "even",))

    def get_selected(self) -> List[Tuple]:
        return [self.tree.item(k)["values"] for k in self.tree.selection()]

    def get_all(self) -> List[Tuple]:
        return [self.tree.item(k)["values"] for k in self.tree.get_children("")]

    def count(self) -> int:
        return len(self.tree.get_children(""))

    def set_headings(self, headings: Tuple[str, ...]) -> None:
        for col, head in zip(self.cols, headings):
            old = self.tree.heading(col)["text"].rstrip(" \u25b2\u25bc")
            arrow = self.tree.heading(col)["text"][len(old):]
            self.tree.heading(col, text=head + arrow)

    def remove_selected(self) -> None:
        for k in list(self.tree.selection()):
            self.tree.delete(k)
        self._restripe()




class LoadingOverlay(ctk.CTkFrame):

    def __init__(self, parent: ctk.CTk) -> None:
        super().__init__(parent, fg_color="#1a1a1a", corner_radius=0)
        self.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.lift()

        card = ctk.CTkFrame(self, fg_color="#242424", corner_radius=16)
        card.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(card,
                     text="Starfield Item Codex",
                     font=ctk.CTkFont(size=28, weight="bold")).pack(padx=60, pady=(35, 6))
        ctk.CTkLabel(card,
                     text=t("loading_msg"),
                     font=ctk.CTkFont(size=13),
                     text_color="#aaaaaa").pack(pady=(0, 18))

        self._bar = ctk.CTkProgressBar(card, mode="indeterminate", width=300)
        self._bar.pack(padx=60, pady=(0, 35))
        self._bar.start()

    def dismiss(self) -> None:
        self._bar.stop()
        self.destroy()



## SearchPanel

class SearchPanel(ctk.CTkFrame):
    """Search box, category/subcategory dropdowns, and results treeview. Shared across all tabs."""

    def __init__(self, parent: tk.Widget, app: "StarfieldItemCodexApp", **kw: Any) -> None:
        super().__init__(parent, fg_color="transparent", **kw)
        self.app = app
        self._all_data: Dict[str, List[Item]] = {}
        self._items: List[Item] = []
        self._filtered: List[Item] = []
        self._added_fids: set = set()  # FormIDs currently in the right-side list

        self._search_var = tk.StringVar()
        self._cat_var    = tk.StringVar(value=t("all_cat"))
        self._subcat_var = tk.StringVar(value=t("all_subcat"))

        # which columns the search bar checks (defaults match original behavior)
        self._srch_name      = tk.BooleanVar(value=True)
        self._srch_formid    = tk.BooleanVar(value=True)
        self._srch_editorid  = tk.BooleanVar(value=True)
        self._srch_category  = tk.BooleanVar(value=False)
        self._srch_subcategory = tk.BooleanVar(value=False)
        self._srch_source    = tk.BooleanVar(value=False)

        self._build_ui()

        self._search_var.trace_add("write", lambda *_: self._apply_filter())
        self._cat_var.trace_add("write",    lambda *_: self._on_cat_change())
        self._subcat_var.trace_add("write", lambda *_: self._apply_filter())

    def _build_ui(self) -> None:
        r0 = ctk.CTkFrame(self, fg_color="transparent")
        r0.pack(fill="x", padx=PAD, pady=(PAD, 4))

        self._search_lbl = ctk.CTkLabel(r0, text=t("search"), width=54, anchor="w")
        self._search_lbl.pack(side="left")

        # dropdown to pick which columns the search bar filters on
        self._col_filter_btn = ctk.CTkButton(
            r0, text="\u25BC", width=28, height=28,
            command=self._show_col_filter_menu)
        self._col_filter_btn.pack(side="left", padx=(2, 0))
        self._col_filter_menu = tk.Menu(
            r0, tearoff=0,
            bg="#2b2b2b", fg="#DCE4EE",
            activebackground=TV_SEL, activeforeground="white", bd=0)
        self._col_filter_menu.add_checkbutton(
            label="Name", variable=self._srch_name,
            command=self._apply_filter)
        self._col_filter_menu.add_checkbutton(
            label="FormID", variable=self._srch_formid,
            command=self._apply_filter)
        self._col_filter_menu.add_checkbutton(
            label="EditorID", variable=self._srch_editorid,
            command=self._apply_filter)
        self._col_filter_menu.add_checkbutton(
            label="Category", variable=self._srch_category,
            command=self._apply_filter)
        self._col_filter_menu.add_checkbutton(
            label="SubCategory", variable=self._srch_subcategory,
            command=self._apply_filter)
        self._col_filter_menu.add_checkbutton(
            label="Source", variable=self._srch_source,
            command=self._apply_filter)

        self.search_entry = ctk.CTkEntry(
            r0, textvariable=self._search_var,
            placeholder_text=t("search_ph"))
        self.search_entry.pack(side="left", padx=(4, 4), fill="x", expand=True)
        ctk.CTkButton(r0, text="X", width=30,
                      command=lambda: self._search_var.set("")).pack(side="left")

        r1 = ctk.CTkFrame(self, fg_color="transparent")
        r1.pack(fill="x", padx=PAD, pady=(0, 6))

        self._cat_lbl = ctk.CTkLabel(r1, text=t("category"), width=72, anchor="w")
        self._cat_lbl.pack(side="left")
        self.cat_cb = ctk.CTkComboBox(r1, variable=self._cat_var,
                                       width=190, state="readonly",
                                       command=lambda _: None)
        self.cat_cb.pack(side="left", padx=(4, PAD))

        self._subcat_lbl = ctk.CTkLabel(r1, text=t("subcategory"), width=88, anchor="w")
        self._subcat_lbl.pack(side="left")
        self.subcat_cb = ctk.CTkComboBox(r1, variable=self._subcat_var,
                                          width=190, state="readonly",
                                          command=lambda _: None)
        self.subcat_cb.pack(side="left", padx=(4, 0))

        self.tv = DarkTreeview(self, LOOKUP_COLS,
                                ("", t("col_category"), t("col_subcategory"), t("col_name"),
                                 t("col_formid"), t("col_editorid"), t("col_source")),
                                LOOKUP_WIDTHS, selectmode="extended")
        self.tv.pack(fill="both", expand=True, padx=PAD, pady=(0, 4))

        self._count_lbl = ctk.CTkLabel(self, text="", anchor="w",
                                        text_color="#888888",
                                        font=ctk.CTkFont(size=11))
        self._count_lbl.pack(fill="x", padx=PAD + 2, pady=(0, 4))


    def init_data(self, data: Dict[str, List[Item]], sheet_names: List[str]) -> None:
        self._all_data = data
        self._load_current_lang()

    def _load_current_lang(self) -> None:
        self._items = self._all_data.get(_current_lang, [])
        self._rebuild_cat_cb()
        self._apply_filter()

    def refresh_ui_text(self) -> None:
        self._search_lbl.configure(text=t("search"))
        self._cat_lbl.configure(text=t("category"))
        self._subcat_lbl.configure(text=t("subcategory"))
        self.search_entry.configure(placeholder_text=t("search_ph"))

        self.tv.set_headings(("", t("col_category"), t("col_subcategory"), t("col_name"),
                              t("col_formid"), t("col_editorid"), t("col_source")))
        self._load_current_lang()

    def set_added_fids(self, fids: set) -> None:
        """Update the set of FormIDs that are 'added' and refresh the display."""
        self._added_fids = fids
        self._apply_filter()

    def get_selected_items(self) -> List[Item]:
        sel = self.tv.get_selected()
        fid_map = {it.form_id: it for it in self._filtered}
        result = []
        for row in sel:
            # col 0 is the checkbox indicator, formid is at index 4
            fid = str(row[4]) if len(row) > 4 else ""
            if fid in fid_map:
                result.append(fid_map[fid])
        return result

    def get_visible_items(self) -> List[Item]:
        return list(self._filtered)

    def get_items_by_category(self, category: str) -> List[Item]:
        return [it for it in self._items if it.category == category]

    def get_items_by_subcategory(self, category: str, subcategory: str) -> List[Item]:
        return [it for it in self._items
                if it.category == category and it.subcategory == subcategory]

    def get_all_categories(self) -> List[str]:
        return sorted({it.category for it in self._items if it.category})

    def get_subcategories_for(self, category: str) -> List[str]:
        return sorted({it.subcategory for it in self._items
                       if it.category == category and it.subcategory})


    def _show_col_filter_menu(self) -> None:
        """pop the column-filter checkbutton menu below the arrow button."""
        btn = self._col_filter_btn
        x = btn.winfo_rootx()
        y = btn.winfo_rooty() + btn.winfo_height()
        self._col_filter_menu.tk_popup(x, y)

    def _on_cat_change(self) -> None:
        self._rebuild_subcat_cb()
        self._apply_filter()

    def _rebuild_cat_cb(self) -> None:
        cats = sorted({it.category for it in self._items if it.category})
        self.cat_cb.configure(values=[t("all_cat")] + cats)
        self._cat_var.set(t("all_cat"))

    def _rebuild_subcat_cb(self) -> None:
        cat = self._cat_var.get()
        all_cat_label = t("all_cat")
        pool = (self._items if cat == all_cat_label
                else [it for it in self._items if it.category == cat])
        subs = sorted({it.subcategory for it in pool if it.subcategory})
        self.subcat_cb.configure(values=[t("all_subcat")] + subs)
        self._subcat_var.set(t("all_subcat"))

    def _apply_filter(self) -> None:
        # TODO: cache this instead of rebuilding every filter change
        query  = self._search_var.get().strip().lower()
        cat    = self._cat_var.get()
        subcat = self._subcat_var.get()
        all_cat_label = t("all_cat")
        all_subcat_label = t("all_subcat")
        hardcoded_only = getattr(self.app, '_hardcoded_var', None)

        result = self._items
        if hardcoded_only and hardcoded_only.get():
            result = [it for it in result if it.source not in VARIABLE_ESMS]
        if cat != all_cat_label:
            result = [it for it in result if it.category == cat]
        if subcat != all_subcat_label:
            result = [it for it in result if it.subcategory == subcat]
        if query:
            # check only the columns the user has enabled in the filter menu
            checks = []
            if self._srch_name.get():        checks.append(lambda it: query in it.name.lower())
            if self._srch_formid.get():      checks.append(lambda it: query in it.form_id.lower())
            if self._srch_editorid.get():    checks.append(lambda it: query in it.editor_id.lower())
            if self._srch_category.get():    checks.append(lambda it: query in it.category.lower())
            if self._srch_subcategory.get(): checks.append(lambda it: query in it.subcategory.lower())
            if self._srch_source.get():      checks.append(lambda it: query in it.source.lower())
            if checks:
                result = [it for it in result if any(fn(it) for fn in checks)]
            else:
                # nothing checked = nothing matches
                result = []

        self._filtered = result
        af = self._added_fids
        rows = [("\u2611" if it.form_id in af else "\u2610",
                 it.category, it.subcategory, it.name,
                 it.form_id, it.editor_id, it.source)
                for it in result]
        self.tv.populate(rows, added_fids=af, fid_col_idx=4)

        n = len(result)
        if n == 1:
            self._count_lbl.configure(text=t("one_item_shown"))
        elif n:
            self._count_lbl.configure(text=t("items_shown", n=f"{n:,}"))
        else:
            self._count_lbl.configure(text=t("no_match"))


# --- tab 1: item lookup ---

class LookupTab:
    """Tab 1: full-width search panel with right-click context menu."""

    def __init__(self, frame: ctk.CTkFrame, app: "StarfieldItemCodexApp") -> None:
        self.app = app
        self.frame = frame

        self.panel = SearchPanel(frame, app)
        self.panel.pack(fill="both", expand=True)

        # right-click context menu
        self._ctx_menu = tk.Menu(frame, tearoff=0,
                                  bg="#2b2b2b", fg="#DCE4EE",
                                  activebackground=TV_SEL, activeforeground="white",
                                  bd=0)
        self._rebuild_ctx_menu()

        tree = self.panel.tv.tree
        tree.bind("<Double-1>",      self._on_double_click)
        tree.bind("<Button-3>",      self._on_right_click)
        tree.bind("<Button-2>",      self._on_right_click)

    def _rebuild_ctx_menu(self) -> None:
        self._ctx_menu.delete(0, "end")
        self._ctx_menu.add_command(label=t("ctx_formid"), command=self._ctx_copy_fid)
        self._ctx_menu.add_command(label=t("ctx_console"), command=self._ctx_copy_cmd)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label=t("ctx_row"), command=self._ctx_copy_row)

    def init_data(self, data: Dict[str, List[Item]], names: List[str]) -> None:
        self.panel.init_data(data, names)

    def refresh_ui_text(self) -> None:
        self.panel.refresh_ui_text()
        self._rebuild_ctx_menu()

    def _on_double_click(self, event: tk.Event) -> None:
        sel = self.panel.tv.get_selected()
        if not sel:
            return
        fid = str(sel[0][4])
        if clipboard_set(self.app, fid):
            self.app.set_status(t("copied_fid", fid=fid))
        else:
            self.app.set_status(t("clipboard_err"))

    def _on_right_click(self, event: tk.Event) -> None:
        row = self.panel.tv.tree.identify_row(event.y)
        if row:
            self.panel.tv.tree.selection_set(row)
            try:
                self._ctx_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self._ctx_menu.grab_release()

    def _ctx_copy_fid(self) -> None:
        sel = self.panel.tv.get_selected()
        if not sel:
            return
        fid = str(sel[0][4])
        clipboard_set(self.app, fid)
        self.app.set_status(t("copied_fid", fid=fid))

    def _ctx_copy_cmd(self) -> None:
        sel = self.panel.tv.get_selected()
        if not sel:
            return
        fid = str(sel[0][4])
        cmd = f"player.additem {fid} 1"
        clipboard_set(self.app, cmd)
        self.app.set_status(t("copied_cmd", cmd=cmd))

    def _ctx_copy_row(self) -> None:
        sel = self.panel.tv.get_selected()
        if not sel:
            return
        # skip col 0 (checkbox indicator) when copying
        row_text = "\t".join(str(v) for v in sel[0][1:])
        clipboard_set(self.app, row_text)
        self.app.set_status(t("copied_row"))


# --- tab 2: subcategory builder ---

class SubcategoryBuilderTab:
    """Left search panel + right subcategory builder. Outputs an INI line."""

    def __init__(self, frame: ctk.CTkFrame, app: "StarfieldItemCodexApp") -> None:
        self.app = app
        self.frame = frame
        self._subcat_items: List[Item] = []
        self._subcat_fids: set = set()

        self._build_ui()

    def _build_ui(self) -> None:
        name_row = ctk.CTkFrame(self.frame, fg_color="transparent")
        name_row.pack(fill="x", padx=PAD, pady=(PAD, 4))

        self._name_lbl = ctk.CTkLabel(name_row, text=t("subcat_name"),
                                       font=ctk.CTkFont(weight="bold"))
        self._name_lbl.pack(side="left")
        self._name_var = tk.StringVar()
        self._name_var.trace_add("write", lambda *_: self._on_name_change())
        self._name_entry = ctk.CTkEntry(
            name_row, textvariable=self._name_var, width=320,
            placeholder_text=t("name_ph"),
            border_color="#b22222")  # red border until a name is entered
        self._name_entry.pack(side="left", padx=(8, 0))

        # output area: pack first (bottom) so it always shows
        out_frame = ctk.CTkFrame(self.frame, fg_color="#1e1e1e", corner_radius=8)
        out_frame.pack(side="bottom", fill="x", padx=PAD, pady=(6, PAD))

        preview_row = ctk.CTkFrame(out_frame, fg_color="transparent")
        preview_row.pack(fill="x", padx=PAD, pady=(8, 2))
        self._ini_lbl = ctk.CTkLabel(preview_row, text=t("ini_preview"),
                                      font=ctk.CTkFont(weight="bold"))
        self._ini_lbl.pack(side="left", padx=(0, 6))
        self._preview_var = tk.StringVar(value="")
        self._preview_lbl = ctk.CTkLabel(
            preview_row, textvariable=self._preview_var, anchor="w",
            text_color="#7ec8e3", font=ctk.CTkFont(family="Courier New", size=11))
        self._preview_lbl.pack(side="left", fill="x", expand=True)

        self._copy_btn = ctk.CTkButton(
            out_frame, text=t("copy_subcat"), width=300, height=36,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._copy_ini_line, state="disabled")
        self._copy_btn.pack(pady=(2, 8))

        # main split: pack after output so it fills remaining space
        split = ctk.CTkFrame(self.frame, fg_color="transparent")
        split.pack(fill="both", expand=True, padx=PAD, pady=(4, 0))
        split.grid_columnconfigure(0, weight=3)
        split.grid_columnconfigure(1, weight=0)
        split.grid_columnconfigure(2, weight=2)
        split.grid_rowconfigure(0, weight=1)

        self.panel = SearchPanel(split, self.app)
        self.panel.grid(row=0, column=0, sticky="nsew")

        mid = ctk.CTkFrame(split, fg_color="transparent", width=130)
        mid.grid(row=0, column=1, sticky="ns", padx=6)
        mid.grid_propagate(False)
        self._mid_frame = mid
        self._build_mid_buttons(mid)

        right = ctk.CTkFrame(split, fg_color="#242424", corner_radius=8)
        right.grid(row=0, column=2, sticky="nsew")
        self._right_frame = right
        self._build_right_panel(right)

        self.panel.tv.tree.bind("<Double-1>", self._search_double_click)
        self._update_preview()

    def _build_mid_buttons(self, parent: ctk.CTkFrame) -> None:
        ctk.CTkLabel(parent, text="", height=40).pack()

        self._btn_add_sel = ctk.CTkButton(parent, text=t("add_selected") + " ->",
                                           command=self._add_selected, width=120)
        self._btn_add_sel.pack(pady=4)
        self._btn_add_vis = ctk.CTkButton(parent, text=t("add_all_vis") + " ->",
                                           command=self._add_all_visible, width=120)
        self._btn_add_vis.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._ent_cat_lbl = ctk.CTkLabel(parent, text=t("entire_cat"),
                                          font=ctk.CTkFont(size=11), text_color="#aaaaaa")
        self._ent_cat_lbl.pack()
        self._add_cat_var = tk.StringVar()
        self._add_cat_cb = ctk.CTkComboBox(parent, variable=self._add_cat_var,
                                            width=120, state="readonly",
                                            command=self._on_add_cat_change)
        self._add_cat_cb.pack(pady=(2, 0))
        self._btn_add_cat = ctk.CTkButton(parent, text=t("add_category") + " ->",
                                           command=self._add_entire_category, width=120)
        self._btn_add_cat.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=4)

        self._ent_subcat_lbl = ctk.CTkLabel(parent, text=t("entire_subcat"),
                                             font=ctk.CTkFont(size=11), text_color="#aaaaaa")
        self._ent_subcat_lbl.pack()
        self._add_subcat_var = tk.StringVar()
        self._add_subcat_cb = ctk.CTkComboBox(parent, variable=self._add_subcat_var,
                                               width=120, state="readonly",
                                               command=lambda _: None)
        self._add_subcat_cb.pack(pady=(2, 0))
        self._btn_add_subcat = ctk.CTkButton(parent, text=t("add_subcat") + " ->",
                                              command=self._add_entire_subcategory, width=120)
        self._btn_add_subcat.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._btn_remove = ctk.CTkButton(parent, text="<- " + t("remove_sel"),
                                          command=self._remove_selected, width=120,
                                          fg_color="#8B1A1A", hover_color="#b22222")
        self._btn_remove.pack(pady=4)
        self._btn_clear = ctk.CTkButton(parent, text=t("clear_all"),
                                         command=self._clear_all, width=120,
                                         fg_color="#5a1e1e", hover_color="#7a2a2a")
        self._btn_clear.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._btn_copy_batch = ctk.CTkButton(
            parent, text=t("copy_from_batch"),
            command=self._copy_from_batch, width=120,
            fg_color="#2a4a6b", hover_color="#3a5a7b")
        self._btn_copy_batch.pack(pady=4)

    def _copy_from_batch(self) -> None:
        batch_items = self.app.batch_tab._batch_items
        if not batch_items:
            self.app.set_status(t("nothing_to_copy"))
            return
        added = self._add_items(batch_items)
        if added:
            self.app.set_status(t("copied_from_batch", n=added))

    def _paste_existing_subcategory(self) -> None:
        """open a popup for pasting existing subcategory lines from an ini."""
        dlg = ctk.CTkToplevel(self.app)
        dlg.title(t("paste_dlg_title"))
        dlg.resizable(True, True)
        dlg.transient(self.app)
        dlg.grab_set()

        # center the dialog over the main window
        dlg_w, dlg_h = 600, 280
        app_x = self.app.winfo_x()
        app_y = self.app.winfo_y()
        app_w = self.app.winfo_width()
        app_h = self.app.winfo_height()
        x = app_x + (app_w - dlg_w) // 2
        y = app_y + (app_h - dlg_h) // 2
        dlg.geometry(f"{dlg_w}x{dlg_h}+{x}+{y}")

        ctk.CTkLabel(dlg, text=t("paste_dlg_instr"), anchor="w",
                     justify="left").pack(fill="x", padx=12, pady=(12, 4))

        textbox = ctk.CTkTextbox(dlg, width=560, height=160,
                                  font=ctk.CTkFont(family="Courier New", size=12))
        textbox.pack(fill="both", expand=True, padx=12, pady=4)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill="x", padx=12, pady=(4, 12))

        def on_add() -> None:
            raw = textbox.get("1.0", "end").strip()
            if not raw:
                self.app.set_status(t("paste_empty"))
                dlg.destroy()
                return
            self._parse_and_import_subcategory(raw)
            dlg.destroy()

        ctk.CTkButton(btn_row, text=t("paste_dlg_add"), width=100,
                      command=on_add).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text=t("paste_dlg_cancel"), width=100,
                      fg_color="#5a1e1e", hover_color="#7a2a2a",
                      command=dlg.destroy).pack(side="left")

        # focus the textbox so user can paste immediately
        textbox.focus_set()

    def _parse_and_import_subcategory(self, raw: str) -> None:
        """parse pasted ini lines and add their formids to the builder list."""
        # build a lookup from the current language's items for fast formid resolution
        fid_lookup: Dict[str, Item] = {}
        items = self.panel._all_data.get(_current_lang, [])
        for it in items:
            fid_lookup[it.form_id.lower()] = it

        parsed_name: str | None = None
        all_fids: List[str] = []  # ordered, preserves paste order
        seen: set = set()

        for line in raw.splitlines():
            line = line.strip()
            if not line or line.startswith(";") or line.startswith("#"):
                continue  # skip blanks, ini comments

            # split on first '=' to get name and formid list
            if "=" in line:
                name_part, fid_part = line.split("=", 1)
                name_part = name_part.strip()
                fid_part = fid_part.strip()
                if parsed_name is None and name_part:
                    parsed_name = name_part
            else:
                fid_part = line

            # split formids on comma
            for raw_fid in fid_part.split(","):
                fid = raw_fid.strip()
                if not fid:
                    continue
                key = fid.lower()
                if key not in seen:
                    seen.add(key)
                    all_fids.append(fid)

        if not all_fids:
            self.app.set_status(t("paste_no_fids"))
            return

        # set the subcategory name from the first parsed line
        if parsed_name:
            self._name_var.set(parsed_name)

        # resolve each formid to a known item or create an unknown placeholder
        to_add: List[Item] = []
        unknown_count = 0
        for fid in all_fids:
            known = fid_lookup.get(fid.lower())
            if known:
                to_add.append(known)
            else:
                # unknown formid, probably a modded item
                to_add.append(Item(
                    category="",
                    subcategory="",
                    name=f"Unknown ({fid})",
                    form_id=fid,
                    editor_id="",
                    source="",
                ))
                unknown_count += 1

        added = self._add_items(to_add)
        self.app.set_status(t("paste_imported", n=added, u=unknown_count))

    def _on_add_cat_change(self, _val: str = "") -> None:
        cat = self._add_cat_var.get()
        if cat:
            subs = self.panel.get_subcategories_for(cat)
            self._add_subcat_cb.configure(values=subs)
            if subs:
                self._add_subcat_var.set(subs[0])
            else:
                self._add_subcat_var.set("")

    def _build_right_panel(self, parent: ctk.CTkFrame) -> None:
        # paste button sits above the header row
        self._btn_paste = ctk.CTkButton(
            parent, text=t("paste_existing"),
            command=self._paste_existing_subcategory, width=220,
            fg_color="#2a4a6b", hover_color="#3a5a7b")
        self._btn_paste.pack(padx=PAD, pady=(PAD, 2))

        hdr = ctk.CTkFrame(parent, fg_color="transparent")
        hdr.pack(fill="x", padx=PAD, pady=(2, 4))
        self._your_subcat_lbl = ctk.CTkLabel(hdr, text=t("your_subcat"),
                                              font=ctk.CTkFont(weight="bold"))
        self._your_subcat_lbl.pack(side="left")
        self._count_badge = ctk.CTkLabel(hdr, text=t("n_items", n=0),
                                          text_color="#aaaaaa",
                                          font=ctk.CTkFont(size=11))
        self._count_badge.pack(side="right")

        self.bldr_tv = DarkTreeview(parent, BLDR_COLS,
                                     (t("col_name"), t("col_formid")),
                                     BLDR_WIDTHS, selectmode="extended")
        self.bldr_tv.pack(fill="both", expand=True, padx=PAD, pady=(0, PAD))
        self.bldr_tv.tree.bind("<Double-1>", self._builder_double_click)


    def init_data(self, data: Dict[str, List[Item]], names: List[str]) -> None:
        self.panel.init_data(data, names)
        self._refresh_add_cat_dropdowns()

    def has_unsaved_work(self) -> bool:
        return len(self._subcat_items) > 0

    def refresh_ui_text(self) -> None:
        self.panel.refresh_ui_text()
        self._name_lbl.configure(text=t("subcat_name"))
        self._name_entry.configure(placeholder_text=t("name_ph"))
        self._ini_lbl.configure(text=t("ini_preview"))
        self._copy_btn.configure(text=t("copy_subcat"))
        self._btn_add_sel.configure(text=t("add_selected") + " ->")
        self._btn_add_vis.configure(text=t("add_all_vis") + " ->")
        self._ent_cat_lbl.configure(text=t("entire_cat"))
        self._btn_add_cat.configure(text=t("add_category") + " ->")
        self._ent_subcat_lbl.configure(text=t("entire_subcat"))
        self._btn_add_subcat.configure(text=t("add_subcat") + " ->")
        self._btn_remove.configure(text="<- " + t("remove_sel"))
        self._btn_clear.configure(text=t("clear_all"))
        self._btn_copy_batch.configure(text=t("copy_from_batch"))
        self._btn_paste.configure(text=t("paste_existing"))
        self._your_subcat_lbl.configure(text=t("your_subcat"))
        self.bldr_tv.set_headings((t("col_name"), t("col_formid")))
        self._update_badge()
        self._update_preview()
        self._refresh_add_cat_dropdowns()


    def _sync_added_markers(self) -> None:
        self.panel.set_added_fids(self._subcat_fids)

    def _add_items(self, items: List[Item]) -> int:
        added = 0
        for it in items:
            if it.form_id not in self._subcat_fids:
                self._subcat_items.append(it)
                self._subcat_fids.add(it.form_id)
                self.bldr_tv.append_row((it.name, it.form_id))
                added += 1
        if added:
            self._update_badge()
            self._update_preview()
            self._sync_added_markers()
            self.app.set_status(t("added_to_sub", n=added))
        return added

    def _update_badge(self) -> None:
        n = len(self._subcat_items)
        self._count_badge.configure(
            text=t("one_item") if n == 1 else t("n_items", n=n))

    def _update_preview(self) -> None:
        name = self._name_var.get().strip()
        if not name or not self._subcat_items:
            self._preview_var.set(t("preview_ph"))
            self._copy_btn.configure(state="disabled")
            return
        fids = ",".join(it.form_id for it in self._subcat_items)
        self._preview_var.set(f"{name}={fids}")
        self._copy_btn.configure(state="normal")

    def _on_name_change(self) -> None:
        raw = self._name_var.get()
        sanitized, changed = sanitize_ini_key(raw)
        if changed:
            cursor = self._name_entry.index(tk.INSERT)
            self._name_var.set(sanitized)
            try:
                self._name_entry.icursor(cursor)
            except Exception:
                pass
            self.app.set_status(t("name_autofixed"))

        # red border when empty, default when filled
        if self._name_var.get().strip():
            self._name_entry.configure(border_color="#565B5E")
        else:
            self._name_entry.configure(border_color="#b22222")

        self._update_preview()

    def _refresh_add_cat_dropdowns(self) -> None:
        cats = self.panel.get_all_categories()
        self._add_cat_cb.configure(values=cats)
        if cats:
            self._add_cat_var.set(cats[0])
            subs = self.panel.get_subcategories_for(cats[0])
            self._add_subcat_cb.configure(values=subs)
            if subs:
                self._add_subcat_var.set(subs[0])

    def _add_selected(self) -> None:
        items = self.panel.get_selected_items()
        if not items:
            self.app.set_status(t("no_sel_search"))
            return
        self._add_items(items)

    def _add_all_visible(self) -> None:
        items = self.panel.get_visible_items()
        if not items:
            self.app.set_status(t("no_vis"))
            return
        self._add_items(items)

    def _add_entire_category(self) -> None:
        cat = self._add_cat_var.get()
        if not cat:
            self.app.set_status(t("sel_cat_first"))
            return
        self._add_items(self.panel.get_items_by_category(cat))

    def _add_entire_subcategory(self) -> None:
        cat = self._add_cat_var.get()
        sub = self._add_subcat_var.get()
        if not cat or not sub:
            self.app.set_status(t("sel_cat_sub"))
            return
        self._add_items(self.panel.get_items_by_subcategory(cat, sub))

    def _remove_selected(self) -> None:
        sel = self.bldr_tv.get_selected()
        if not sel:
            self.app.set_status(t("no_sel_sub"))
            return
        sel_fids = {str(row[1]) for row in sel}
        self._subcat_items = [it for it in self._subcat_items
                               if it.form_id not in sel_fids]
        self._subcat_fids -= sel_fids
        self.bldr_tv.populate([(it.name, it.form_id) for it in self._subcat_items])
        self._update_badge()
        self._update_preview()
        self._sync_added_markers()
        self.app.set_status(t("removed_sub", n=len(sel_fids)))

    def _clear_all(self) -> None:
        if not self._subcat_items:
            return
        if not messagebox.askyesno(
                t("dlg_clear_all"),
                t("dlg_clear_sub", n=len(self._subcat_items)),
                parent=self.frame):
            return
        self._subcat_items.clear()
        self._subcat_fids.clear()
        self.bldr_tv.clear()
        self._update_badge()
        self._update_preview()
        self._sync_added_markers()
        self.app.set_status(t("cleared_sub"))

    def _copy_ini_line(self) -> None:
        line = self._preview_var.get()
        if not line or line.startswith("("):
            return
        if clipboard_set(self.app, line):
            self.app.set_status(t("ini_copied"))
        else:
            self.app.set_status(t("clipboard_err"))

    def _search_double_click(self, event: tk.Event) -> None:
        items = self.panel.get_selected_items()
        if not items:
            return
        # double-click toggles: remove if already added, otherwise add
        to_add = [it for it in items if it.form_id not in self._subcat_fids]
        to_remove = [it for it in items if it.form_id in self._subcat_fids]
        if to_remove:
            for it in to_remove:
                self._subcat_items = [x for x in self._subcat_items if x.form_id != it.form_id]
                self._subcat_fids.discard(it.form_id)
            self.bldr_tv.populate([(it.name, it.form_id) for it in self._subcat_items])
            self._update_badge()
            self._update_preview()
            self._sync_added_markers()
            self.app.set_status(t("removed_sub", n=len(to_remove)))
        if to_add:
            self._add_items(to_add)

    def _builder_double_click(self, event: tk.Event) -> None:
        sel = self.bldr_tv.get_selected()
        if not sel:
            return
        fid = str(sel[0][1])
        self._subcat_items = [it for it in self._subcat_items if it.form_id != fid]
        self._subcat_fids.discard(fid)
        self.bldr_tv.populate([(it.name, it.form_id) for it in self._subcat_items])
        self._update_badge()
        self._update_preview()
        self._sync_added_markers()
        self.app.set_status(t("removed_item"))


# tab 3: batch file creator

class BatchCreatorTab:
    """Search panel + batch list. Editable qtys, saves to .txt batch file."""

    def __init__(self, frame: ctk.CTkFrame, app: "StarfieldItemCodexApp") -> None:
        self.app = app
        self.frame = frame
        self._batch_items: List[Item] = []
        self._batch_fids: set = set()
        self._batch_qtys: Dict[str, int] = {}
        self._qty_edit_widget: Optional[tk.Entry] = None

        self._build_ui()

    def _build_ui(self) -> None:
        # output area: pack bottom-first so it always shows
        out_frame = ctk.CTkFrame(self.frame, fg_color="#1e1e1e", corner_radius=8)
        out_frame.pack(side="bottom", fill="x", padx=PAD, pady=(6, PAD))

        name_row = ctk.CTkFrame(out_frame, fg_color="transparent")
        name_row.pack(fill="x", padx=PAD, pady=(8, 4))
        self._batch_name_lbl = ctk.CTkLabel(name_row, text=t("batch_name"),
                                             font=ctk.CTkFont(weight="bold"))
        self._batch_name_lbl.pack(side="left")
        self._batch_name_var = tk.StringVar()
        self._batch_name_var.trace_add("write", lambda *_: self._refresh_output_buttons())
        self._batch_name_entry = ctk.CTkEntry(
            name_row, textvariable=self._batch_name_var, width=250,
            placeholder_text=t("batch_name_ph"))
        self._batch_name_entry.pack(side="left", padx=(8, 0), fill="x", expand=True)

        btn_row = ctk.CTkFrame(out_frame, fg_color="transparent")
        btn_row.pack(fill="x", padx=PAD, pady=(0, 8))

        self._save_btn = ctk.CTkButton(
            btn_row, text=t("save_batch"), width=180, height=34,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._save_batch, state="disabled")
        self._save_btn.pack(side="left", padx=(0, 8))

        self._copy_cmds_btn = ctk.CTkButton(
            btn_row, text=t("copy_commands"), width=160, height=34,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._copy_commands, state="disabled")
        self._copy_cmds_btn.pack(side="left", padx=(0, 8))

        self._out_info_lbl = ctk.CTkLabel(
            btn_row, text=t("add_to_enable"),
            text_color="#666666", anchor="w")
        self._out_info_lbl.pack(side="left", padx=6)

        # main split: pack after output so it fills remaining space
        split = ctk.CTkFrame(self.frame, fg_color="transparent")
        split.pack(fill="both", expand=True, padx=PAD, pady=(PAD, 0))
        split.grid_columnconfigure(0, weight=3)
        split.grid_columnconfigure(1, weight=0)
        split.grid_columnconfigure(2, weight=2)
        split.grid_rowconfigure(0, weight=1)

        self.panel = SearchPanel(split, self.app)
        self.panel.grid(row=0, column=0, sticky="nsew")

        mid = ctk.CTkFrame(split, fg_color="transparent", width=130)
        mid.grid(row=0, column=1, sticky="ns", padx=6)
        mid.grid_propagate(False)
        self._mid_frame = mid
        self._build_mid_buttons(mid)

        right = ctk.CTkFrame(split, fg_color="#242424", corner_radius=8)
        right.grid(row=0, column=2, sticky="nsew")
        self._right_frame = right
        self._build_right_panel(right)

        self.panel.tv.tree.bind("<Double-1>", self._search_double_click)

    def _build_mid_buttons(self, parent: ctk.CTkFrame) -> None:
        ctk.CTkLabel(parent, text="", height=40).pack()

        self._btn_add_sel = ctk.CTkButton(parent, text=t("add_selected") + " ->",
                                           command=self._add_selected, width=120)
        self._btn_add_sel.pack(pady=4)
        self._btn_add_vis = ctk.CTkButton(parent, text=t("add_all_vis") + " ->",
                                           command=self._add_all_visible, width=120)
        self._btn_add_vis.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._ent_cat_lbl = ctk.CTkLabel(parent, text=t("entire_cat"),
                                          font=ctk.CTkFont(size=11), text_color="#aaaaaa")
        self._ent_cat_lbl.pack()
        self._add_cat_var = tk.StringVar()
        self._add_cat_cb = ctk.CTkComboBox(parent, variable=self._add_cat_var,
                                            width=120, state="readonly",
                                            command=self._on_add_cat_change)
        self._add_cat_cb.pack(pady=(2, 0))
        self._btn_add_cat = ctk.CTkButton(parent, text=t("add_category") + " ->",
                                           command=self._add_entire_category, width=120)
        self._btn_add_cat.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=4)

        self._ent_subcat_lbl = ctk.CTkLabel(parent, text=t("entire_subcat"),
                                             font=ctk.CTkFont(size=11), text_color="#aaaaaa")
        self._ent_subcat_lbl.pack()
        self._add_subcat_var = tk.StringVar()
        self._add_subcat_cb = ctk.CTkComboBox(parent, variable=self._add_subcat_var,
                                               width=120, state="readonly",
                                               command=lambda _: None)
        self._add_subcat_cb.pack(pady=(2, 0))
        self._btn_add_subcat = ctk.CTkButton(parent, text=t("add_subcat") + " ->",
                                              command=self._add_entire_subcategory, width=120)
        self._btn_add_subcat.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._btn_remove = ctk.CTkButton(parent, text="<- " + t("remove_sel"),
                                          command=self._remove_selected, width=120,
                                          fg_color="#8B1A1A", hover_color="#b22222")
        self._btn_remove.pack(pady=4)
        self._btn_clear = ctk.CTkButton(parent, text=t("clear_all"),
                                         command=self._clear_all, width=120,
                                         fg_color="#5a1e1e", hover_color="#7a2a2a")
        self._btn_clear.pack(pady=4)

        ctk.CTkFrame(parent, fg_color="#444444", height=1).pack(fill="x", pady=8)

        self._btn_copy_bldr = ctk.CTkButton(
            parent, text=t("copy_from_bldr"),
            command=self._copy_from_builder, width=120,
            fg_color="#2a4a6b", hover_color="#3a5a7b")
        self._btn_copy_bldr.pack(pady=4)

    def _copy_from_builder(self) -> None:
        bldr_items = self.app.subcat_tab._subcat_items
        if not bldr_items:
            self.app.set_status(t("nothing_to_copy"))
            return
        added = self._add_items(bldr_items)
        if added:
            self.app.set_status(t("copied_from_bldr", n=added))

    def _on_add_cat_change(self, _val: str = "") -> None:
        cat = self._add_cat_var.get()
        if cat:
            subs = self.panel.get_subcategories_for(cat)
            self._add_subcat_cb.configure(values=subs)
            if subs:
                self._add_subcat_var.set(subs[0])
            else:
                self._add_subcat_var.set("")

    def _build_right_panel(self, parent: ctk.CTkFrame) -> None:
        hdr = ctk.CTkFrame(parent, fg_color="transparent")
        hdr.pack(fill="x", padx=PAD, pady=(PAD, 4))
        self._batch_items_lbl = ctk.CTkLabel(hdr, text=t("batch_items"),
                                              font=ctk.CTkFont(weight="bold"))
        self._batch_items_lbl.pack(side="left")
        self._count_badge = ctk.CTkLabel(hdr, text=t("n_items", n=0),
                                          text_color="#aaaaaa",
                                          font=ctk.CTkFont(size=11))
        self._count_badge.pack(side="right")

        self.batch_tv = DarkTreeview(parent, BATCH_COLS,
                                      (t("col_name"), t("col_formid"), t("col_qty")),
                                      BATCH_WIDTHS, selectmode="extended")
        self.batch_tv.pack(fill="both", expand=True, padx=PAD, pady=(0, 4))
        self.batch_tv.tree.bind("<ButtonRelease-1>", self._on_batch_click)
        self.batch_tv.tree.bind("<Double-1>", self._batch_double_click)

        qty_row = ctk.CTkFrame(parent, fg_color="transparent")
        qty_row.pack(fill="x", padx=PAD, pady=(0, PAD))
        self._set_qty_lbl = ctk.CTkLabel(qty_row, text=t("set_all_qty"))
        self._set_qty_lbl.pack(side="left")
        self._set_all_qty_var = tk.StringVar(value="1")
        ctk.CTkEntry(qty_row, textvariable=self._set_all_qty_var,
                     width=60, justify="center").pack(side="left", padx=(6, 6))
        self._btn_apply_qty = ctk.CTkButton(qty_row, text=t("apply"), width=70,
                                             command=self._set_all_qty)
        self._btn_apply_qty.pack(side="left")

        # set qty for selected items only
        self._set_sel_qty_lbl = ctk.CTkLabel(qty_row, text=t("set_sel_qty"))
        self._set_sel_qty_lbl.pack(side="left", padx=(20, 0))
        self._set_sel_qty_var = tk.StringVar(value="1")
        ctk.CTkEntry(qty_row, textvariable=self._set_sel_qty_var,
                     width=60, justify="center").pack(side="left", padx=(6, 6))
        self._btn_apply_sel_qty = ctk.CTkButton(qty_row, text=t("apply"), width=70,
                                                command=self._set_selected_qty)
        self._btn_apply_sel_qty.pack(side="left")


    def init_data(self, data: Dict[str, List[Item]], names: List[str]) -> None:
        self.panel.init_data(data, names)
        self._refresh_add_dropdowns()

    def has_unsaved_work(self) -> bool:
        return len(self._batch_items) > 0

    def refresh_ui_text(self) -> None:
        self.panel.refresh_ui_text()
        self._batch_name_lbl.configure(text=t("batch_name"))
        self._batch_name_entry.configure(placeholder_text=t("batch_name_ph"))
        self._save_btn.configure(text=t("save_batch"))
        self._copy_cmds_btn.configure(text=t("copy_commands"))
        self._btn_add_sel.configure(text=t("add_selected") + " ->")
        self._btn_add_vis.configure(text=t("add_all_vis") + " ->")
        self._ent_cat_lbl.configure(text=t("entire_cat"))
        self._btn_add_cat.configure(text=t("add_category") + " ->")
        self._ent_subcat_lbl.configure(text=t("entire_subcat"))
        self._btn_add_subcat.configure(text=t("add_subcat") + " ->")
        self._btn_remove.configure(text="<- " + t("remove_sel"))
        self._btn_clear.configure(text=t("clear_all"))
        self._btn_copy_bldr.configure(text=t("copy_from_bldr"))
        self._batch_items_lbl.configure(text=t("batch_items"))
        self._set_qty_lbl.configure(text=t("set_all_qty"))
        self._btn_apply_qty.configure(text=t("apply"))
        self._set_sel_qty_lbl.configure(text=t("set_sel_qty"))
        self._btn_apply_sel_qty.configure(text=t("apply"))
        self.batch_tv.set_headings((t("col_name"), t("col_formid"), t("col_qty")))
        self._update_badge()
        self._refresh_output_buttons()
        self._refresh_add_dropdowns()

    # inline qty cell editing

    def _on_batch_click(self, event: tk.Event) -> None:
        region = self.batch_tv.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.batch_tv.tree.identify_column(event.x)
        if col != "#3":
            return
        row_id = self.batch_tv.tree.identify_row(event.y)
        if not row_id:
            return
        self._start_qty_edit(row_id)

    def _start_qty_edit(self, row_id: str) -> None:
        if self._qty_edit_widget is not None:
            try:
                self._qty_edit_widget.destroy()
            except Exception:
                pass
            self._qty_edit_widget = None

        try:
            bbox = self.batch_tv.tree.bbox(row_id, "qty")
        except Exception:
            return
        if not bbox:
            return

        x, y, w, h = bbox
        current = self.batch_tv.tree.set(row_id, "qty")
        var = tk.StringVar(value=str(current))

        entry = tk.Entry(
            self.batch_tv.tree, textvariable=var,
            bg=TV_SEL, fg="white", insertbackground="white",
            relief="flat", bd=2, justify="center",
            font=("TkDefaultFont", 10))
        entry.place(x=x, y=y, width=max(w, 55), height=h)
        entry.select_range(0, "end")
        entry.focus_set()
        self._qty_edit_widget = entry

        def commit(_event: Any = None) -> None:
            if self._qty_edit_widget is not entry:
                return
            try:
                v = max(1, int(var.get()))
            except (ValueError, TypeError):
                v = 1
            fid = str(self.batch_tv.tree.set(row_id, "form_id"))
            self.batch_tv.tree.set(row_id, "qty", str(v))
            self._batch_qtys[fid] = v
            entry.destroy()
            self._qty_edit_widget = None

        def cancel(_event: Any = None) -> None:
            if self._qty_edit_widget is entry:
                entry.destroy()
                self._qty_edit_widget = None

        entry.bind("<Return>",   commit)
        entry.bind("<Tab>",      commit)
        entry.bind("<Escape>",   cancel)
        entry.bind("<FocusOut>", commit)


    def _sync_added_markers(self) -> None:
        self.panel.set_added_fids(self._batch_fids)

    def _add_items(self, items: List[Item]) -> int:
        added = 0
        for it in items:
            if it.form_id not in self._batch_fids:
                self._batch_items.append(it)
                self._batch_fids.add(it.form_id)
                self._batch_qtys[it.form_id] = 1
                self.batch_tv.append_row((it.name, it.form_id, "1"))
                added += 1
        if added:
            self._update_badge()
            self._refresh_output_buttons()
            self._sync_added_markers()
            self.app.set_status(t("added_to_batch", n=added))
        return added

    def _update_badge(self) -> None:
        n = len(self._batch_items)
        self._count_badge.configure(
            text=t("one_item") if n == 1 else t("n_items", n=n))

    def _refresh_output_buttons(self) -> None:
        has_items = len(self._batch_items) > 0
        has_name = bool(self._batch_name_var.get().strip())
        # save needs both name + items; copy just needs items
        self._save_btn.configure(state="normal" if (has_items and has_name) else "disabled")
        self._copy_cmds_btn.configure(state="normal" if has_items else "disabled")
        self._out_info_lbl.configure(
            text=t("n_ready", n=len(self._batch_items)) if has_items else t("add_to_enable"))

    def _refresh_add_dropdowns(self) -> None:
        cats = self.panel.get_all_categories()
        self._add_cat_cb.configure(values=cats)
        if cats:
            self._add_cat_var.set(cats[0])
            subs = self.panel.get_subcategories_for(cats[0])
            self._add_subcat_cb.configure(values=subs)
            if subs:
                self._add_subcat_var.set(subs[0])

    def _add_selected(self) -> None:
        items = self.panel.get_selected_items()
        if not items:
            self.app.set_status(t("no_sel_search"))
            return
        self._add_items(items)

    def _add_all_visible(self) -> None:
        items = self.panel.get_visible_items()
        if not items:
            self.app.set_status(t("no_vis"))
            return
        self._add_items(items)

    def _add_entire_category(self) -> None:
        cat = self._add_cat_var.get()
        if not cat:
            self.app.set_status(t("sel_cat_first"))
            return
        self._add_items(self.panel.get_items_by_category(cat))

    def _add_entire_subcategory(self) -> None:
        cat = self._add_cat_var.get()
        sub = self._add_subcat_var.get()
        if not cat or not sub:
            self.app.set_status(t("sel_cat_sub"))
            return
        self._add_items(self.panel.get_items_by_subcategory(cat, sub))

    def _remove_selected(self) -> None:
        sel = self.batch_tv.get_selected()
        if not sel:
            self.app.set_status(t("no_sel_batch"))
            return
        sel_fids = {str(row[1]) for row in sel}
        self._batch_items = [it for it in self._batch_items
                              if it.form_id not in sel_fids]
        self._batch_fids -= sel_fids
        for fid in sel_fids:
            self._batch_qtys.pop(fid, None)
        self.batch_tv.populate(
            [(it.name, it.form_id, str(self._batch_qtys.get(it.form_id, 1)))
             for it in self._batch_items])
        self._update_badge()
        self._refresh_output_buttons()
        self._sync_added_markers()
        self.app.set_status(t("removed_batch", n=len(sel_fids)))

    def _clear_all(self) -> None:
        if not self._batch_items:
            return
        if not messagebox.askyesno(
                t("dlg_clear_all"),
                t("dlg_clear_batch", n=len(self._batch_items)),
                parent=self.frame):
            return
        self._batch_items.clear()
        self._batch_fids.clear()
        self._batch_qtys.clear()
        self.batch_tv.clear()
        self._update_badge()
        self._refresh_output_buttons()
        self._sync_added_markers()
        self.app.set_status(t("cleared_batch"))

    def _set_all_qty(self) -> None:
        try:
            v = max(1, int(self._set_all_qty_var.get()))
        except (ValueError, TypeError):
            self.app.set_status(t("invalid_qty"))
            return
        self._set_all_qty_var.set(str(v))
        for it in self._batch_items:
            self._batch_qtys[it.form_id] = v
        for row_id in self.batch_tv.tree.get_children(""):
            self.batch_tv.tree.set(row_id, "qty", str(v))
        self.app.set_status(t("qty_set", n=v))

    def _set_selected_qty(self) -> None:
        """set qty only for highlighted items in the batch list."""
        sel = self.batch_tv.tree.selection()
        if not sel:
            self.app.set_status(t("no_sel_for_qty"))
            return
        try:
            v = max(1, int(self._set_sel_qty_var.get()))
        except (ValueError, TypeError):
            self.app.set_status(t("invalid_qty"))
            return
        self._set_sel_qty_var.set(str(v))
        for row_id in sel:
            fid = str(self.batch_tv.tree.set(row_id, "form_id"))
            self._batch_qtys[fid] = v
            self.batch_tv.tree.set(row_id, "qty", str(v))
        self.app.set_status(t("qty_set_sel", c=len(sel), n=v))

    def _search_double_click(self, event: tk.Event) -> None:
        items = self.panel.get_selected_items()
        if not items:
            return
        # double-click toggles: remove if already added, otherwise add
        to_add = [it for it in items if it.form_id not in self._batch_fids]
        to_remove = [it for it in items if it.form_id in self._batch_fids]
        if to_remove:
            remove_fids = {it.form_id for it in to_remove}
            self._batch_items = [x for x in self._batch_items if x.form_id not in remove_fids]
            self._batch_fids -= remove_fids
            for fid in remove_fids:
                self._batch_qtys.pop(fid, None)
            self.batch_tv.populate(
                [(it.name, it.form_id, str(self._batch_qtys.get(it.form_id, 1)))
                 for it in self._batch_items])
            self._update_badge()
            self._refresh_output_buttons()
            self._sync_added_markers()
            self.app.set_status(t("removed_batch", n=len(remove_fids)))
        if to_add:
            self._add_items(to_add)

    def _batch_double_click(self, event: tk.Event) -> None:
        """double-click on right panel removes the item."""
        sel = self.batch_tv.get_selected()
        if not sel:
            return
        fid = str(sel[0][1])
        self._batch_items = [it for it in self._batch_items if it.form_id != fid]
        self._batch_fids.discard(fid)
        self._batch_qtys.pop(fid, None)
        self.batch_tv.populate(
            [(it.name, it.form_id, str(self._batch_qtys.get(it.form_id, 1)))
             for it in self._batch_items])
        self._update_badge()
        self._refresh_output_buttons()
        self._sync_added_markers()
        self.app.set_status(t("removed_batch", n=1))

    # output

    def _build_batch_text(self) -> str:
        lines = [
            "; ==========================================",
            "; Generated by Starfield Item Codex",
            "; Usage: Place in Starfield Data folder",
            ";        Open console (~) and type: bat filename",
            "; ==========================================",
            "",
        ]
        for it in self._batch_items:
            qty = self._batch_qtys.get(it.form_id, 1)
            lines.append(f"; {it.name}")
            lines.append(f"player.additem {it.form_id} {qty}")
            lines.append("")
        return "\n".join(lines)

    def _save_batch(self) -> None:
        if not self._batch_items:
            return
        raw_name = self._batch_name_var.get().strip()
        if not raw_name:
            return
        safe_name = raw_name.replace(" ", "_")
        safe_name = re.sub(r'[<>:"/\\|?*]', '', safe_name)
        if not safe_name:
            safe_name = "AstralUI_batch"
        path = filedialog.asksaveasfilename(
            title=t("save_batch"),
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"{safe_name}.txt",
            parent=self.frame)
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self._build_batch_text())
            self.app.set_status(t("batch_saved", f=os.path.basename(path)))
        except OSError as exc:
            messagebox.showerror(t("dlg_save_err"),
                                  t("dlg_write_fail", e=str(exc)),
                                  parent=self.frame)

    def _copy_commands(self) -> None:
        if not self._batch_items:
            return
        lines = []
        for it in self._batch_items:
            qty = self._batch_qtys.get(it.form_id, 1)
            lines.append(f"player.additem {it.form_id} {qty}")
        text = "\n".join(lines)
        if clipboard_set(self.app, text):
            self.app.set_status(t("cmds_copied", n=len(lines)))
        else:
            self.app.set_status(t("clipboard_err"))


## StarfieldItemCodexApp ##

class StarfieldItemCodexApp(ctk.CTk):
    """Main window. Three tabs, global language selector, status bar at bottom."""

    def __init__(self) -> None:
        super().__init__()

        self.title(APP_TITLE)
        self.geometry(f"{APP_W}x{APP_H}")
        self.minsize(MIN_W, MIN_H)
        self._all_data: Dict[str, List[Item]] = {}
        self._sheet_names: List[str] = []

        setup_treeview_style()

        # status bar
        self._status_var = tk.StringVar(value=t("starting"))
        status_bar = ctk.CTkFrame(self, fg_color="#1a1a1a", height=28, corner_radius=0)
        status_bar.pack(side="bottom", fill="x")
        status_bar.pack_propagate(False)
        ctk.CTkLabel(status_bar, textvariable=self._status_var,
                     anchor="w", text_color="#aaaaaa",
                     font=ctk.CTkFont(size=11)).pack(side="left", padx=12, pady=4)
        self._status_ts = ctk.CTkLabel(status_bar, text="",
                                        anchor="e", text_color="#666666",
                                        font=ctk.CTkFont(size=11))
        self._status_ts.pack(side="right", padx=12, pady=4)

        # top bar
        top_frame = ctk.CTkFrame(self, fg_color="#1e1e1e", corner_radius=6)
        top_frame.pack(fill="x", padx=PAD, pady=(PAD, 4))

        # language selector row
        lang_row = ctk.CTkFrame(top_frame, fg_color="transparent")
        lang_row.pack(fill="x", padx=12, pady=(8, 2))
        self._lang_lbl = ctk.CTkLabel(lang_row, text=t("language"), anchor="w",
                                       font=ctk.CTkFont(weight="bold"))
        self._lang_lbl.pack(side="left")
        self._lang_var = tk.StringVar(value="English")
        self._lang_cb = ctk.CTkComboBox(lang_row, variable=self._lang_var,
                                         width=280, state="readonly",
                                         command=self._on_lang_display_change)
        self._lang_cb.pack(side="left", padx=(8, 0))

        self._hardcoded_var = tk.BooleanVar(value=True)
        self._hardcoded_cb = ctk.CTkCheckBox(
            lang_row, text=t("hardcoded_only"),
            variable=self._hardcoded_var,
            font=ctk.CTkFont(size=13))
        self._hardcoded_cb.pack(side="left", padx=(24, 0))
        self._hardcoded_var.trace_add("write", lambda *_: self._on_hardcoded_change())

        # description lines
        self._desc_labels: List[ctk.CTkLabel] = []
        for key, emoji in [("desc_lookup", "🔍"), ("desc_builder", "🏗"),
                            ("desc_batch", "🎮")]:
            lbl = ctk.CTkLabel(top_frame, text=f"{emoji}  {t(key)}", anchor="w",
                               text_color="#999999", font=ctk.CTkFont(size=12))
            lbl.pack(anchor="w", padx=12, pady=(2, 0))
            self._desc_labels.append(lbl)
        ctk.CTkFrame(top_frame, fg_color="transparent", height=4).pack()

        self._desc_keys = ["desc_lookup", "desc_batch", "desc_builder"]
        self._desc_emojis = ["🔍", "🎮", "🏗"]

        # tabs
        self._prev_tab = "🔍 Search"
        self._tab_guard = False

        # use english tab names as internal keys
        self._tab_keys = list(TAB_NAMES_EN.values())  # ['\U0001f50d Lookup', ...]
        self._tab_id_map = dict(zip(self._tab_keys, TAB_IDS))  # display -> id

        self.tabview = ctk.CTkTabview(
            self, command=self._on_tab_change,
            segmented_button_fg_color="#2b2b2b",
            segmented_button_selected_color="#3B8ED0",
            segmented_button_selected_hover_color="#3696D9",
            segmented_button_unselected_color="#3a3a3a",
            segmented_button_unselected_hover_color="#505050",
            text_color="#DCE4EE",
            text_color_disabled="#888888")
        self.tabview.pack(fill="both", expand=True, padx=PAD, pady=(4, 4))

        try:
            seg = self.tabview._segmented_button
            seg.configure(font=ctk.CTkFont(size=15, weight="bold"), height=40)
        except Exception:
            pass

        for name in self._tab_keys:
            self.tabview.add(name)

        self.lookup_tab  = LookupTab(self.tabview.tab(self._tab_keys[0]), self)
        self.batch_tab   = BatchCreatorTab(self.tabview.tab(self._tab_keys[1]), self)
        self.subcat_tab  = SubcategoryBuilderTab(self.tabview.tab(self._tab_keys[2]), self)

        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._overlay = LoadingOverlay(self)
        self._start_load()


    def _on_hardcoded_change(self) -> None:
        if hasattr(self, 'lookup_tab'):
            self.lookup_tab.panel._apply_filter()
        if hasattr(self, 'subcat_tab'):
            self.subcat_tab.panel._apply_filter()
        if hasattr(self, 'batch_tab'):
            self.batch_tab.panel._apply_filter()

    def _on_lang_display_change(self, display_name: str) -> None:
        sheet_name = LANG_FROM_DISPLAY.get(display_name, display_name)
        self._on_lang_change(sheet_name)

    def _on_lang_change(self, new_lang: str) -> None:
        global _current_lang
        if new_lang not in self._all_data:
            return
        _current_lang = new_lang
        # keep the combobox showing the display name
        self._lang_var.set(LANG_DISPLAY_NAMES.get(new_lang, new_lang))

        for lbl, key, emoji in zip(self._desc_labels, self._desc_keys, self._desc_emojis):
            lbl.configure(text=f"{emoji}  {t(key)}")

        self._lang_lbl.configure(text=t("language"))
        self._hardcoded_cb.configure(text=t("hardcoded_only"))

        self._update_tab_labels()

        self.lookup_tab.refresh_ui_text()
        self.subcat_tab.refresh_ui_text()
        self.batch_tab.refresh_ui_text()

        self.set_status(t("loaded_status",
                          total=f"{len(self._all_data.get(new_lang, [])):,}",
                          langs=len(self._sheet_names)))

    def _update_tab_labels(self) -> None:
        # HACK: CTk doesn't expose tab renaming cleanly so we poke at internals
        names = TAB_NAMES.get(_current_lang, TAB_NAMES_EN)
        try:
            seg = self.tabview._segmented_button
            current_val = self.tabview.get()
            new_labels = [names[tid] for tid in TAB_IDS]
            old_labels = self._tab_keys
            for i, btn_key in enumerate(old_labels):
                if hasattr(seg, '_buttons_dict') and btn_key in seg._buttons_dict:
                    seg._buttons_dict[btn_key].configure(text=new_labels[i])
        except Exception:
            pass  # graceful fallback if CTk internals change


    def _start_load(self) -> None:
        path = find_xlsx()
        if not path:
            self.after(100, self._ask_for_xlsx)
        else:
            threading.Thread(target=self._load_thread, args=(path,),
                             daemon=True).start()

    def _ask_for_xlsx(self) -> None:
        messagebox.showinfo(
            t("dlg_not_found"),
            t("dlg_not_found_msg", name=XLSX_NAME),
            parent=self)
        path = filedialog.askopenfilename(
            title=f"Open {XLSX_NAME}",
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
            parent=self)
        if not path:
            messagebox.showerror(t("dlg_no_file"), t("dlg_no_file_msg"), parent=self)
            self.destroy()
            return
        threading.Thread(target=self._load_thread, args=(path,),
                         daemon=True).start()

    def _load_thread(self, path: str) -> None:
        data, names, error = load_xlsx(path)
        self.after(0, self._on_data_loaded, data, names, error)

    def _on_data_loaded(self, data: Dict[str, List[Item]],
                         names: List[str], error: Optional[str]) -> None:
        self._overlay.dismiss()

        if error:
            messagebox.showerror(t("dlg_load_err"), error, parent=self)
            self.set_status(t("load_failed", e=error.splitlines()[0]))
            return

        self._all_data = data
        self._sheet_names = names

        display_names = [LANG_DISPLAY_NAMES.get(n, n) for n in names]
        self._lang_cb.configure(values=display_names)
        if names:
            self._lang_var.set(LANG_DISPLAY_NAMES.get(names[0], names[0]))

        total = sum(len(v) for v in data.values())
        self.set_status(t("loaded_status", total=f"{total:,}", langs=len(names)))

        self.lookup_tab.init_data(data, names)
        self.subcat_tab.init_data(data, names)
        self.batch_tab.init_data(data, names)


    def set_status(self, message: str) -> None:
        self._status_var.set(message)
        ts = datetime.now().strftime("%H:%M:%S")
        self._status_ts.configure(text=ts)


    def _on_tab_change(self) -> None:
        if self._tab_guard:
            return
        self._prev_tab = self.tabview.get()

    def _on_close(self) -> None:
        self.destroy()



def main() -> None:
    app = StarfieldItemCodexApp()
    app.mainloop()


if __name__ == "__main__":
    main()
